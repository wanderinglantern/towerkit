"""Generic styled-table writer: SOI styling for arbitrary sectioned tables."""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from towerkit.render.table_xlsx import TableColumn, TableSection, write_table
from towerkit.theme import load_theme

COLS = (
    TableColumn("Item", 30.0),
    TableColumn("Amount", 12.0, number_format='"$"#,##0.00', align="right"),
)
SECTIONS = (
    TableSection("Section A", (("first", 100), ("second", 200)), total=300),
    TableSection(None, (("loose row", 5),)),
)


@pytest.fixture()
def theme():
    return load_theme(None)


def test_headers_sections_and_total(theme, tmp_path: Path):
    path = write_table(COLS, SECTIONS, title="T", theme=theme, out_path=tmp_path / "t.xlsx")
    ws = load_workbook(path).active
    assert [c.value for c in ws[1]] == ["Item", "Amount"]
    assert ws["A2"].value == "Section A"      # label row
    assert ws["B2"].value == 300              # total in last column
    assert ws["A3"].value == "first"
    assert ws["A5"].value == "loose row"      # unlabeled section has no label row
    assert ws.freeze_panes == "A2"


def test_two_runs_byte_identical(theme, tmp_path: Path):
    a = write_table(COLS, SECTIONS, title="T", theme=theme, out_path=tmp_path / "a.xlsx")
    b = write_table(COLS, SECTIONS, title="T", theme=theme, out_path=tmp_path / "b.xlsx")
    assert a.read_bytes() == b.read_bytes()


def test_row_height_hook(theme, tmp_path: Path):
    path = write_table(
        COLS, SECTIONS, title="T", theme=theme, out_path=tmp_path / "h.xlsx",
        row_height=lambda values: 44.0,
    )
    ws = load_workbook(path).active
    assert ws.row_dimensions[3].height == 44.0


def test_title_sanitized_for_illegal_sheet_chars(theme, tmp_path: Path):
    # openpyxl raises ValueError for / \ ? * [ ] : in a sheet title —
    # write_table is the authority, so no caller can crash on a name like
    # "A/B Partners". This must not raise, and the surviving title must
    # still be recognizably the client name.
    path = write_table(
        COLS, SECTIONS, title="A/B [test]: weird?", theme=theme,
        out_path=tmp_path / "sanitized.xlsx",
    )
    wb = load_workbook(path)
    title = wb.active.title
    assert title == "A B test weird"
    for char in "/\\?*[]:":
        assert char not in title
    assert len(title) <= 31


def test_two_sheet_workbook_single_normalize(theme, tmp_path: Path):
    from openpyxl import Workbook

    from towerkit.render.table_xlsx import finalize_workbook, render_table_sheet

    def build(path: Path) -> Path:
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "First"
        render_table_sheet(ws, COLS, SECTIONS, theme=theme)
        second = wb.create_sheet("Second")
        render_table_sheet(second, COLS, SECTIONS, theme=theme)
        return finalize_workbook(wb, path)

    a = build(tmp_path / "a.xlsx")
    wb = load_workbook(a)
    assert wb.sheetnames == ["First", "Second"]
    assert [c.value for c in wb["Second"][1]] == ["Item", "Amount"]  # styled body on sheet 2
    assert wb["Second"].freeze_panes == "A2"
    # one finalize → still deterministic
    assert a.read_bytes() == build(tmp_path / "b.xlsx").read_bytes()


def test_finalize_pins_modified_timestamp(theme, tmp_path: Path):
    import zipfile

    from openpyxl import Workbook

    from towerkit.render.table_xlsx import finalize_workbook

    path = finalize_workbook(Workbook(), tmp_path / "p.xlsx")
    with zipfile.ZipFile(path) as z:
        core = z.read("docProps/core.xml").decode()
    assert "1980-01-01T00:00:00Z</dcterms:modified>" in core


def test_new_workbook_starts_multi_sheet_composition(theme, tmp_path: Path):
    # bookkit cannot name openpyxl (convention-tested) nor re-import Workbook
    # from this namespace (its strict mypy: no_implicit_reexport) — this
    # factory is how convention-bound consumers start the multi-sheet flow.
    from towerkit.render.table_xlsx import (
        finalize_workbook,
        new_workbook,
        render_table_sheet,
        sanitize_sheet_title,
    )

    wb = new_workbook()
    ws = wb.active
    ws.title = sanitize_sheet_title("One")
    render_table_sheet(ws, COLS, SECTIONS, theme=theme)
    ws2 = wb.create_sheet(sanitize_sheet_title("Two"))
    render_table_sheet(ws2, COLS, SECTIONS, theme=theme)
    path = finalize_workbook(wb, tmp_path / "multi.xlsx")
    assert load_workbook(path).sheetnames == ["One", "Two"]
