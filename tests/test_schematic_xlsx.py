"""Schematic worksheet: quantizer first (pure), then cell content (Task 4)."""

import re
import xml.etree.ElementTree as ET
import zipfile
from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from test_soi import make_program

from towerkit import __version__
from towerkit.layout import build_layout
from towerkit.model import Layer, Line, Participant, Period, Placement, Program, Retention
from towerkit.render.labels import participant_label
from towerkit.render.schematic_xlsx import (
    AXIS_COL,
    AXIS_WIDTH,
    CANVAS_WIDTH_UNITS,
    FIRST_GRID_COL,
    FIRST_GRID_ROW,
    GROUP_ROW,
    HEADER_TWO_LINE_HEIGHT,
    LINE_ROW,
    SPACER_UNITS,
    TOTAL_ROWS,
    SchematicOverlapError,
    _label_spans,
    add_schematic_sheet,
    label_row_floor,
    line_column_slots,
    quantize_boundaries,
    sheet_rows,
    x_boundaries,
    y_boundaries,
)
from towerkit.render.table_xlsx import _argb, finalize_workbook
from towerkit.theme import load_theme
from towerkit.validate import validate_program

REPO = Path(__file__).parent.parent
_SML_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"


def _mini_program() -> Program:
    """One line, a $1M primary under a $25M excess split 60/40."""
    return Program(
        insured="T", program="T", placement=Placement.BOUND,
        period=Period(start=date(2026, 1, 1), end=date(2027, 1, 1)),
        lines=[Line(id="gl", name="General Liability")],
        layers=[
            Layer(id="p", name="Primary", applies_to=["gl"], attach=0,
                  limit=1_000_000,
                  participants=[Participant(carrier="A", share_bps=10_000)]),
            Layer(id="x", name="Excess", applies_to=["gl"], attach=1_000_000,
                  limit=25_000_000,
                  participants=[Participant(carrier="B", share_bps=6_000),
                                Participant(carrier="C", share_bps=4_000)]),
        ],
    )


class TestQuantizeBoundaries:
    def test_full_coverage_and_proportionality(self) -> None:
        rows = quantize_boundaries([0.0, 0.1, 1.0], total_rows=100)
        assert rows[0.0] == 0 and rows[0.1] == 10 and rows[1.0] == 100

    def test_min_one_row_floor_is_strict_monotonicity(self) -> None:
        rows = quantize_boundaries([0.0, 0.001, 0.002, 1.0], total_rows=100)
        assert rows[0.001] == 1 and rows[0.002] == 2

    def test_negative_retention_band_shares_the_grid(self) -> None:
        rows = quantize_boundaries([-0.18, 0.0, 1.0], total_rows=100)
        assert rows[-0.18] == 0 and rows[1.0] == 100

    def test_adjacent_spans_tile_exactly(self) -> None:
        rows = quantize_boundaries([0.0, 0.3, 1.0])
        below_top, below_bottom = sheet_rows(rows, 0.0, 0.3)
        above_top, above_bottom = sheet_rows(rows, 0.3, 1.0)
        assert above_top == FIRST_GRID_ROW            # y=1.0 is the first grid row
        assert above_bottom + 1 == below_top          # shared boundary, no gap
        assert below_bottom == FIRST_GRID_ROW + rows[1.0] - 1


class TestLayoutBoundaries:
    def test_x_boundaries_include_share_splits(self) -> None:
        xs = x_boundaries(build_layout(_mini_program()))
        assert xs == (0.0, 0.6, 1.0)                  # 60/40 split at exactly 0.6

    def test_y_boundaries_include_every_attachment_and_zero(self) -> None:
        layout = build_layout(_mini_program())
        ys = y_boundaries(layout)
        assert ys[0] == 0.0 and ys[-1] == 1.0
        assert layout.ymap.y(1_000_000) in ys         # the primary/excess boundary


@pytest.fixture()
def marsh():
    return load_theme(REPO / "themes" / "marsh.json")


@pytest.fixture()
def program():
    return make_program()


def _write_schematic(program, theme, path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "SOI"  # stand-in first sheet; Task 5 provides the real one
    add_schematic_sheet(wb, program, theme)
    return finalize_workbook(wb, path)


def _grid(program):
    """Reuses the SAME production pure functions `add_schematic_sheet`
    calls (`_label_spans`, `quantize_boundaries`, `x_boundaries`,
    `line_column_slots`) rather than reimplementing the column mapping —
    Task 11 injects a schematic-only spacer column at every joined line-of-
    cover boundary (see `line_column_slots`'s docstring), so a naive
    `enumerate(xs)` mapping here would silently diverge from what's
    actually on the sheet. Returns (layout, rows, col_of, col_of_close):
    `col_of[x]` is right for an OPENING edge (rect.x0), `col_of_close[x]`
    for a CLOSING edge (rect.x1: c1 = col_of_close[x] - 1) — see
    `line_column_slots`."""
    layout = build_layout(program)
    label_spans = _label_spans(layout, program)
    rows = quantize_boundaries(y_boundaries(layout), label_spans=label_spans)
    xs = x_boundaries(layout)
    col_of, col_of_close = line_column_slots(layout, xs)
    return layout, rows, col_of, col_of_close


def _fill_hex(xlsx_path: Path, sheet_xml: str, cell_ref: str) -> str | None:
    """ARGB fill of one cell, read from the saved XML (test_soi_xlsx.py's
    technique: openpyxl won't report styles on non-anchor merged cells)."""
    with zipfile.ZipFile(xlsx_path) as z:
        styles = ET.fromstring(z.read("xl/styles.xml"))
        sheet = ET.fromstring(z.read(f"xl/worksheets/{sheet_xml}"))
    fills = []
    for fill in styles.find(f"{_SML_NS}fills"):
        pattern = fill.find(f"{_SML_NS}patternFill")
        fg = pattern.find(f"{_SML_NS}fgColor") if pattern is not None else None
        fills.append(fg.get("rgb") if fg is not None else None)
    fill_ids = [int(xf.get("fillId", "0")) for xf in styles.find(f"{_SML_NS}cellXfs")]
    for cell in sheet.iter(f"{_SML_NS}c"):
        if cell.get("r") == cell_ref:
            return fills[fill_ids[int(cell.get("s", "0"))]]
    return None


def _bottom_border_hex(xlsx_path: Path, sheet_xml: str, cell_ref: str) -> str | None:
    """ARGB colour of a cell's bottom border edge, read from the saved XML —
    same styles.xml-reading technique as `_fill_hex`, needed for the same
    reason: openpyxl won't report style on non-anchor merged cells once the
    file is reloaded."""
    with zipfile.ZipFile(xlsx_path) as z:
        styles = ET.fromstring(z.read("xl/styles.xml"))
        sheet = ET.fromstring(z.read(f"xl/worksheets/{sheet_xml}"))
    borders = []
    for border in styles.find(f"{_SML_NS}borders"):
        bottom = border.find(f"{_SML_NS}bottom")
        colour = bottom.find(f"{_SML_NS}color") if bottom is not None else None
        borders.append(colour.get("rgb") if colour is not None else None)
    border_ids = [int(xf.get("borderId", "0")) for xf in styles.find(f"{_SML_NS}cellXfs")]
    for cell in sheet.iter(f"{_SML_NS}c"):
        if cell.get("r") == cell_ref:
            return borders[border_ids[int(cell.get("s", "0"))]]
    return None


def _border_hex(xlsx_path: Path, sheet_xml: str, cell_ref: str, side: str) -> str | None:
    """ARGB colour of one cell's border SIDE ('left'/'right'/'top'/
    'bottom'), read from the saved XML — same styles.xml-reading technique
    as `_bottom_border_hex`, generalized: Task 11's axis lines (left/
    bottom) and attachment gridlines (top) need every side, not just
    bottom."""
    with zipfile.ZipFile(xlsx_path) as z:
        styles = ET.fromstring(z.read("xl/styles.xml"))
        sheet = ET.fromstring(z.read(f"xl/worksheets/{sheet_xml}"))
    borders = []
    for border in styles.find(f"{_SML_NS}borders"):
        edge = border.find(f"{_SML_NS}{side}")
        colour = edge.find(f"{_SML_NS}color") if edge is not None else None
        borders.append(colour.get("rgb") if colour is not None else None)
    border_ids = [int(xf.get("borderId", "0")) for xf in styles.find(f"{_SML_NS}cellXfs")]
    for cell in sheet.iter(f"{_SML_NS}c"):
        if cell.get("r") == cell_ref:
            return borders[border_ids[int(cell.get("s", "0"))]]
    return None


class TestSchematicSheet:
    def test_sheet_appended_and_named(self, program, marsh, tmp_path):
        wb = load_workbook(_write_schematic(program, marsh, tmp_path / "s.xlsx"))
        assert wb.sheetnames == ["SOI", "Casualty Schematic"]

    def test_merged_ranges_never_overlap(self, program, marsh, tmp_path):
        ws = load_workbook(_write_schematic(program, marsh, tmp_path / "s.xlsx"))[
            "Casualty Schematic"
        ]
        bounds = [r.bounds for r in ws.merged_cells.ranges]  # (c0, r0, c1, r1)
        for i, a in enumerate(bounds):
            for b in bounds[i + 1 :]:
                disjoint = a[2] < b[0] or b[2] < a[0] or a[3] < b[1] or b[3] < a[1]
                assert disjoint, f"overlapping merges: {a} vs {b}"

    def test_block_merge_lands_exactly_where_the_layout_says(
        self, program, marsh, tmp_path
    ):
        path = _write_schematic(program, marsh, tmp_path / "s.xlsx")
        ws = load_workbook(path)["Casualty Schematic"]
        layout, rows, col_of, col_of_close = _grid(program)
        zenith = next(
            b for b in layout.participants
            if b.carrier == "Zenith" and b.layer_id == "gl-primary"
        )
        rect = zenith.rects[0]
        r0, r1 = sheet_rows(rows, rect.y0, rect.y1)
        c0, c1 = col_of[rect.x0], col_of_close[rect.x1] - 1
        ref = f"{get_column_letter(c0)}{r0}:{get_column_letter(c1)}{r1}"
        assert ref in {str(r) for r in ws.merged_cells.ranges}
        anchor = ws[f"{get_column_letter(c0)}{r0}"]
        assert "Zenith 100%" in anchor.value
        assert "Primary — $1M" in anchor.value      # heading rides the widest block
        assert "$50K" in anchor.value               # premium share, like the graphic
        # first-appearance palette assignment, same order as the graphic:
        assert anchor.fill.fgColor.rgb == "FF000F47"  # marsh carrierPalette[0]

    def test_merged_interior_cells_carry_the_fill(self, program, marsh, tmp_path):
        path = _write_schematic(program, marsh, tmp_path / "s.xlsx")
        layout, rows, col_of, _ = _grid(program)
        gamma_blk = next(b for b in layout.participants if b.carrier == "Gamma")
        rect = gamma_blk.rects[0]
        r0, _ = sheet_rows(rows, rect.y0, rect.y1)
        c0 = col_of[rect.x0]
        interior = f"{get_column_letter(c0)}{r0 + 1}"  # prop-primary spans many rows
        assert _fill_hex(path, "sheet2.xml", interior) == "FF82BAFF"  # Gamma = palette[3]

    def test_proportions_boundaries_and_stacking(self, program, marsh, tmp_path):
        layout, rows, _, _ = _grid(program)

        def span(layer_id: str) -> tuple[int, int]:
            ly = next(lyr for lyr in layout.layers if lyr.layer_id == layer_id)
            return sheet_rows(rows, ly.y0, ly.y1)

        p_top, p_bottom = span("gl-primary")
        x_top, x_bottom = span("gl-x1")
        assert x_bottom + 1 == p_top                       # stacked flush, no gap
        prop_top, prop_bottom = span("prop-primary")
        assert (prop_bottom - prop_top) > (p_bottom - p_top)  # $10M towers over $1M

    def test_pending_retention_axis_and_lines(self, program, marsh, tmp_path):
        path = _write_schematic(program, marsh, tmp_path / "s.xlsx")
        ws = load_workbook(path)["Casualty Schematic"]
        values = {
            c.value for row in ws.iter_rows() for c in row if c.value is not None
        }
        assert any("To be placed" in v for v in values)     # al-primary is pending
        assert "SIR $250K" in values                        # retention label
        assert any(v == "$1M" for v in values)              # axis boundary label
        assert "General Liability" in values                # line header
        assert any(v.startswith("Casualty — Limit") for v in values)  # group band

    def test_a_buffer_reads_uninsured_not_to_be_placed(self, marsh, tmp_path):
        """Fix round 3 (Grant, 2026-08-21): same error as the vector export
        — a buffer has no participants, so `is_pending` was true for it
        exactly as for a genuinely pending layer, and this worksheet printed
        "To be placed" over a band the broker deliberately left uninsured.
        `_participant_lines` (the single seam both call sites in this module
        share) now takes `buffer=`, routed through the same
        `layer_heading`/`unplaced_label` render/web.py and mpl_program.py
        use, so no renderer can say a different thing about a buffer than
        the others do.

        Uses a standalone program, not the shared `program` fixture: none of
        the golden-hash fixtures in this repo contain a buffer layer, so
        this assertion needs its own."""
        # NAME DELIBERATELY CONTAINS NEITHER "uninsured" NOR "buffer": the
        # fix appends those words itself, and a layer already named
        # "Uninsured Band" would let a positive assertion pass off the
        # LAYER'S OWN NAME rather than anything unplaced_label/layer_heading
        # actually produced — the same class of vacuous assertion this task
        # has already caught twice (round 1's is-buffer class name, round
        # 2's title attribute).
        program = Program(
            insured="T", program="T", placement=Placement.BOUND,
            period=Period(start=date(2026, 1, 1), end=date(2027, 1, 1)),
            lines=[Line(id="gl", name="General Liability")],
            layers=[
                Layer(id="p", name="Primary", applies_to=["gl"], attach=0,
                      limit=5_000_000,
                      participants=[Participant(carrier="A", share_bps=10_000)]),
                Layer(id="b", name="Second Excess", applies_to=["gl"],
                      attach=5_000_000, limit=5_000_000, buffer=True),
            ],
        )
        wb = Workbook()
        wb.remove(wb.active)
        add_schematic_sheet(wb, program, marsh)
        ws = wb.worksheets[0]
        values = {
            c.value for row in ws.iter_rows() for c in row if c.value is not None
        }
        assert any("Uninsured" in v for v in values)
        assert any("buffer" in v.lower() for v in values)
        assert not any("To be placed" in v for v in values)

    def test_retention_fill_is_the_typed_theme_fill(self, program, marsh, tmp_path):
        path = _write_schematic(program, marsh, tmp_path / "s.xlsx")
        layout, rows, col_of, _ = _grid(program)
        sir = next(r for r in layout.retentions if r.type == "sir")
        rect = sir.rects[0]
        r0, _ = sheet_rows(rows, rect.y0, rect.y1)
        anchor = f"{get_column_letter(col_of[rect.x0])}{r0}"
        assert _fill_hex(path, "sheet2.xml", anchor) == "FFFFF3DA"  # marsh sir fill

    def test_two_writes_byte_identical(self, program, marsh, tmp_path):
        a = _write_schematic(program, marsh, tmp_path / "a.xlsx")
        b = _write_schematic(program, marsh, tmp_path / "b.xlsx")
        assert a.read_bytes() == b.read_bytes()

    def test_canvas_fills_the_full_working_width(self, program, marsh, tmp_path):
        """Grant's Excel review (2026-08-13): the tower occupied roughly
        columns A-R and looked cramped. The sheet's total width — the fixed
        axis column plus every proportional tower column (Task 11: plus the
        fixed-width line spacers between them, see SPACER_UNITS) — must
        fill the full working width, not whatever a fixed per-unit rate
        happens to add up to.

        Summed over the FULL contiguous column range (not just
        `x_boundaries()`'s own entries): Task 11 injects extra spacer
        columns between joined lines of cover that have no `x_boundaries`
        entry of their own (see `line_column_slots`), so summing only
        `xs[:-1]` would silently miss them."""
        path = _write_schematic(program, marsh, tmp_path / "s.xlsx")
        ws = load_workbook(path)["Casualty Schematic"]
        layout, _, _, col_of_close = _grid(program)
        xs = x_boundaries(layout)
        last_col = col_of_close[xs[-1]] - 1
        tower_total = sum(
            ws.column_dimensions[get_column_letter(c)].width or 0.0
            for c in range(FIRST_GRID_COL, last_col + 1)
        )
        total = AXIS_WIDTH + tower_total
        # Literal floor, independent of CANVAS_WIDTH_UNITS: pins Grant's
        # actual intent (a total canvas of ~230-250 Excel width units,
        # narrowed from ~300-330 in the 2026-08-13 follow-up review — "each
        # line of cover slightly less wide") rather than merely checking
        # the code's arithmetic is self-consistent. A regression back to
        # the old cramped ~45-unit canvas (the pre-Task-10 fixed
        # 10-units-per-line-width rate on this 3-column fixture) must fail
        # this assertion.
        assert total >= 200.0
        # Secondary check: the module constant is actually honored, not just
        # some other value >= the floor.
        assert abs(total - CANVAS_WIDTH_UNITS) < 1.0


# Task 9: visual parity polish (Grant's Excel review) — theme colours, header
# wrap, narrow-merge fitting, the row floor, and the group-band border fix.


def _filler_lines(n: int) -> list[Line]:
    """Undrawn sibling columns (no layer touches them) that only exist to
    inflate a fixture's total tower span. Task 10 widened the canvas to
    fill CANVAS_WIDTH_UNITS regardless of line count (Grant's review
    2026-08-13), so a single-column fixture no longer renders narrow — it
    gets the WHOLE canvas to itself. These fillers put the column under
    test back in the many-lines-sharing-one-canvas position a real cramped
    tower is actually in, without adding anything for the test to read."""
    return [Line(id=f"filler{i}", name=f"Filler {i}") for i in range(n)]


def _long_header_program() -> Program:
    """A narrow column whose line name cannot fit one wrapped line at the
    column's width — line headers must never clip (requirement 2). Ten
    filler siblings (see `_filler_lines`) keep this column's share of the
    canvas narrow post-Task-10; without them the lone column would claim
    the full ~290-unit canvas and never need to wrap."""
    return Program(
        insured="T", program="T", placement=Placement.BOUND,
        period=Period(start=date(2026, 1, 1), end=date(2027, 1, 1)),
        lines=[Line(id="x", name="Excess Casualty Following Form Umbrella"),
               *_filler_lines(10)],
        layers=[
            Layer(id="p", name="Primary", applies_to=["x"], attach=0, limit=1_000_000,
                  participants=[Participant(carrier="A", share_bps=10_000)]),
        ],
    )


def _narrow_split_program() -> Program:
    """Grant's Property tower case: a 3-way ~33.33% split on one line column
    among many (below the threshold for a 2- or 3-line label, requirement
    3). Nineteen filler siblings (see `_filler_lines`) keep the Property
    column's share of the canvas narrow post-Task-10 — same technique and
    same reason as `_long_header_program`."""
    return Program(
        insured="T", program="Property", placement=Placement.BOUND,
        period=Period(start=date(2026, 1, 1), end=date(2027, 1, 1)),
        lines=[Line(id="prop", name="Property"), *_filler_lines(19)],
        layers=[
            Layer(id="p", name="Primary", applies_to=["prop"], attach=0,
                  limit=9_000_000, premium=90_000,
                  participants=[
                      Participant(carrier="Alpha", share_bps=3_333),
                      Participant(carrier="Beta", share_bps=3_333),
                      Participant(carrier="Gamma", share_bps=3_334),
                  ]),
        ],
    )


def _three_column_group_program() -> Program:
    """Three lines sharing one group bucket — a GROUP_ROW band spanning 3
    columns, whose interior (non-edge) column is where the pre-fix merge-
    before-style bug loses the bottom accent border (requirement 5)."""
    return Program(
        insured="T", program="T", placement=Placement.BOUND,
        period=Period(start=date(2026, 1, 1), end=date(2027, 1, 1)),
        lines=[
            Line(id="gl", name="General Liability", group="Casualty"),
            Line(id="al", name="Auto Liability", group="Casualty"),
            Line(id="el", name="Employers Liability", group="Casualty"),
        ],
        layers=[
            Layer(id="p", name="Primary", applies_to=["gl", "al", "el"], attach=0,
                  limit=3_000_000,
                  participants=[Participant(carrier="A", share_bps=10_000)]),
        ],
    )


class TestVisualPolish:
    def test_every_participant_fill_matches_the_shared_colour_authority(
        self, program, marsh, tmp_path
    ):
        """Theme parity: every participant block's fill must equal marsh's
        theme.carrier_colours assignment — the SAME authority the graphic
        (mpl_program) draws from — not a hardcoded hex and not a colour
        re-derived independently by the schematic."""
        path = _write_schematic(program, marsh, tmp_path / "s.xlsx")
        layout, rows, col_of, _ = _grid(program)
        expected = marsh.carrier_colours(program.carriers())
        checked = 0
        for block in layout.participants:
            if block.carrier is None:
                continue
            anchor_rect = max(block.rects, key=lambda r: r.width)
            r0, _ = sheet_rows(rows, anchor_rect.y0, anchor_rect.y1)
            ref = f"{get_column_letter(col_of[anchor_rect.x0])}{r0}"
            actual = _fill_hex(path, "sheet2.xml", ref)
            assert actual == _argb(expected[block.carrier])
            checked += 1
        assert checked >= 5  # make_program's fixture covers 5 participant blocks

    def test_line_header_wraps_get_a_two_line_row_height(self, marsh, tmp_path):
        program = _long_header_program()
        path = _write_schematic(program, marsh, tmp_path / "s.xlsx")
        ws = load_workbook(path)["T Schematic"]
        anchor = ws.cell(row=LINE_ROW, column=FIRST_GRID_COL)
        assert anchor.alignment.wrap_text is True
        assert ws.row_dimensions[LINE_ROW].height >= HEADER_TWO_LINE_HEIGHT

    def test_narrow_share_split_merges_fall_back_to_compact_labels(
        self, marsh, tmp_path
    ):
        program = _narrow_split_program()
        path = _write_schematic(program, marsh, tmp_path / "s.xlsx")
        ws = load_workbook(path)["Property Schematic"]
        layout, rows, col_of, _ = _grid(program)
        expected = {
            participant_label(carrier, share_bps)
            for carrier, share_bps in (("Alpha", 3_333), ("Beta", 3_333), ("Gamma", 3_334))
        }
        texts = set()
        for block in layout.participants:
            anchor_rect = max(block.rects, key=lambda r: r.width)
            r0, _ = sheet_rows(rows, anchor_rect.y0, anchor_rect.y1)
            c0 = col_of[anchor_rect.x0]
            cell = ws.cell(row=r0, column=c0)
            texts.add(cell.value)
            assert cell.alignment.shrink_to_fit is True
            assert cell.alignment.wrap_text is not True
        # no block below the threshold shows more than the compact
        # "Carrier share%" line — the layer heading never survives it
        assert texts == expected
        assert all("\n" not in (t or "") for t in texts)
        assert not any(" — " in (t or "") for t in texts)

    def test_thin_labeled_layer_gets_the_row_floor(self, tmp_path):
        """Isolates the quantization floor from gamma's own compression
        assist: a $2M primary under a realistic $100M tower is ALREADY kept
        comfortably visible by gamma compression alone (span 20 rows under
        DEFAULT_GAMMA) — this fixture uses a far more extreme ratio ($2K
        under a ~$1B tower) so the pre-fix span genuinely collapses to a
        single row, proving the floor itself, not gamma, does the work."""
        theme = load_theme(None)
        program = Program(
            insured="T", program="T", placement=Placement.BOUND,
            period=Period(start=date(2026, 1, 1), end=date(2027, 1, 1)),
            lines=[Line(id="gl", name="General Liability")],
            layers=[
                Layer(id="p", name="Primary", applies_to=["gl"], attach=0,
                      limit=2_000, premium=100,
                      participants=[Participant(carrier="A", share_bps=10_000)]),
                Layer(id="x", name="Excess", applies_to=["gl"], attach=2_000,
                      limit=999_998_000, premium=500_000,
                      participants=[Participant(carrier="B", share_bps=10_000)]),
            ],
        )
        path = _write_schematic(program, theme, tmp_path / "s.xlsx")
        layout, rows, col_of, _ = _grid(program)
        primary = next(ly for ly in layout.layers if ly.layer_id == "p")
        r0, r1 = sheet_rows(rows, primary.y0, primary.y1)
        assert r1 - r0 + 1 >= 2

        # not narrow in x (a full 1.0-wide column): the label stays wrapped,
        # not shrunk — the row floor and the narrow-merge fix are orthogonal
        ws = load_workbook(path)["T Schematic"]
        anchor = ws.cell(row=r0, column=col_of[layout.columns[0].ex0])
        assert anchor.alignment.wrap_text is True
        assert anchor.alignment.shrink_to_fit is not True

    def test_row_floor_reaches_a_stepped_follows_underlying_run(self, tmp_path):
        """A follows-underlying layer's LayerBlock.y0/y1 is only nominally
        the layer's full height — its stepped-bottom runs (what's actually
        drawn, one per underlying column) can each be a thin fraction of
        that. Two GL columns with a near-total underlying give the umbrella
        a wide-but-y-thin stepped run there; a third Property column with a
        small underlying gives it a narrow-but-y-tall run elsewhere — so the
        wide (anchor-winning) run is also the thin one, exactly the case a
        nominal-layer-span floor would miss and a per-rendered-rect floor
        catches."""
        theme = load_theme(None)
        program = Program(
            insured="T", program="T", placement=Placement.BOUND,
            period=Period(start=date(2026, 1, 1), end=date(2027, 1, 1)),
            lines=[Line(id="gl1", name="GL1"), Line(id="gl2", name="GL2"),
                   Line(id="prop", name="Prop")],
            layers=[
                Layer(id="gl1-p", name="GL1 Primary", applies_to=["gl1"], attach=0,
                      limit=999_998_000,
                      participants=[Participant(carrier="A", share_bps=10_000)]),
                Layer(id="gl2-p", name="GL2 Primary", applies_to=["gl2"], attach=0,
                      limit=999_998_000,
                      participants=[Participant(carrier="A", share_bps=10_000)]),
                Layer(id="prop-p", name="Prop Primary", applies_to=["prop"], attach=0,
                      limit=1_000_000,
                      participants=[Participant(carrier="B", share_bps=10_000)]),
                Layer(id="umb", name="Umbrella", applies_to=["gl1", "gl2", "prop"],
                      attach=0, limit=1_000_000_000, follows_underlying=True,
                      participants=[Participant(carrier="C", share_bps=10_000)]),
            ],
        )
        path = _write_schematic(program, theme, tmp_path / "s.xlsx")
        layout, rows, col_of, _ = _grid(program)
        block = next(b for b in layout.participants if b.layer_id == "umb")
        anchor_rect = max(block.rects, key=lambda r: r.width)
        assert anchor_rect.x1 - anchor_rect.x0 > 1.125  # the wide GL1+GL2 run won
        r0, r1 = sheet_rows(rows, anchor_rect.y0, anchor_rect.y1)
        assert r1 - r0 + 1 >= 2

        ws = load_workbook(path)["T Schematic"]
        anchor = ws.cell(row=r0, column=col_of[anchor_rect.x0])
        assert anchor.value  # the anchor rect really is the one carrying text

    def test_quantize_boundaries_floors_a_labeled_span_to_two_rows(self):
        """Pure-function proof of the floor mechanism, same style as
        TestQuantizeBoundaries above: without label_spans, y=0.001 gets the
        existing 1-row strict-monotonic floor; with a label_spans entry
        naming (0.0, 0.001, min_rows), it must additionally clear its OWN
        min_rows (Task 11: the floor rides per-span/label-aware now, not a
        single module constant — see `label_row_floor`)."""
        bare = quantize_boundaries([0.0, 0.001, 1.0], total_rows=100)
        assert bare[0.001] - bare[0.0] == 1

        floored = quantize_boundaries(
            [0.0, 0.001, 1.0], total_rows=100, label_spans=[(0.0, 0.001, 2)],
        )
        assert floored[0.001] - floored[0.0] >= 2

    def test_quantize_boundaries_floor_rides_per_span(self):
        """Two label_spans ending at the SAME boundary but with different
        min_rows both apply independently — proves the floor is genuinely
        per-span (Task 11), not a single value shared across every labeled
        span like the pre-Task-11 uniform ROW_FLOOR was."""
        rows = quantize_boundaries(
            [0.0, 0.2, 1.0], total_rows=100,
            label_spans=[(0.0, 0.2, 3), (0.2, 1.0, 9)],
        )
        assert rows[0.2] - rows[0.0] >= 3
        assert rows[1.0] - rows[0.2] >= 9

    def test_group_band_bottom_border_survives_a_three_column_span(self, tmp_path):
        theme = load_theme(None)
        program = _three_column_group_program()
        path = _write_schematic(program, theme, tmp_path / "s.xlsx")
        layout, _, col_of, col_of_close = _grid(program)
        band = next(b for b in layout.groups if b.label == "Casualty")
        c0, c1 = col_of[band.x0], col_of_close[band.x1] - 1
        # A genuine 3-column band: 3 real line columns PLUS the 2 Task-11
        # line spacers injected between them (see `line_column_slots`) —
        # the band's own merge spans straight through both, so its column
        # count grew from the pre-Task-11 2 to 4.
        assert c1 - c0 == 4
        # The interior (non-edge) REAL column — al's own, not one of the
        # spacers either side of it — is where the pre-fix merge-before-
        # style bug lost the bottom accent border.
        al_ex0 = next(c for c in layout.columns if c.line_id == "al").ex0
        interior = col_of[al_ex0]
        ref = f"{get_column_letter(interior)}{GROUP_ROW}"
        assert _bottom_border_hex(path, "sheet2.xml", ref) == _argb(theme.chrome.accent)


# Task 11: polish round 3 (Grant's native-Excel review) — narrower canvas,
# label-aware row floor, real axis lines, freeze panes, line-of-cover
# spacers, print setup, faint attachment gridlines and a version-only
# provenance footer.


def _uneven_towers_program() -> Program:
    """Two ungrouped (so layout.py JOINS them flush — "both ungrouped
    counts as one bucket") lines with different tower heights: "tall" runs
    a primary + excess to $10M, "short" stops at $1M. At the row for the
    $10M boundary (tall-x's own top, also the sheet's topmost grid row),
    "tall" is filled (tall-x) and "short" is genuinely blank — the known
    empty-vs-filled contrast the gridline test (item 7) needs, and the
    joined-line-spacer test (item 5) needs an adjacent pair with no group
    in common."""
    return Program(
        insured="T", program="T", placement=Placement.BOUND,
        period=Period(start=date(2026, 1, 1), end=date(2027, 1, 1)),
        lines=[Line(id="tall", name="Tall"), Line(id="short", name="Short")],
        layers=[
            Layer(id="short-p", name="Primary", applies_to=["short"], attach=0,
                  limit=1_000_000,
                  participants=[Participant(carrier="A", share_bps=10_000)]),
            Layer(id="tall-p", name="Primary", applies_to=["tall"], attach=0,
                  limit=1_000_000,
                  participants=[Participant(carrier="B", share_bps=10_000)]),
            Layer(id="tall-x", name="Excess", applies_to=["tall"], attach=1_000_000,
                  limit=9_000_000,
                  participants=[Participant(carrier="C", share_bps=10_000)]),
        ],
    )


class TestNarrowerCanvasAndLabelAwareFloor:
    """Items 1 and 2: the canvas itself narrowed (covered by the updated
    literal in test_canvas_fills_the_full_working_width above); this class
    covers the label-aware row floor."""

    def test_label_row_floor_scales_with_line_count(self) -> None:
        """Pure-function proof of the derivation: ceil(lines *
        LABEL_LINE_HEIGHT_PT / GRID_ROW_HEIGHT) + 1, literal per line
        count (not re-deriving the formula, just pinning its output for a
        few concrete inputs)."""
        assert label_row_floor(1) == 4
        assert label_row_floor(2) == 7
        assert label_row_floor(3) == 10

    def test_thin_primary_three_line_label_clears_the_label_aware_floor(
        self, tmp_path
    ) -> None:
        """Literal-value proof of the LABEL-AWARE floor in an actual
        render (Grant's second Excel review, item 2: "primary-layer labels
        ... render tight against block edges"). The SAME extreme-ratio
        fixture as test_thin_labeled_layer_gets_the_row_floor above (so
        gamma compression isn't doing the work) — its sole/widest
        participant carries the heading, share AND premium, a genuine
        3-line label. The span must clear 10 rows: a LITERAL pinned here,
        independent of quantize_boundaries' and label_row_floor's own
        arithmetic, per the project's 'never self-referential' rule for
        floor tests."""
        theme = load_theme(None)
        program = Program(
            insured="T", program="T", placement=Placement.BOUND,
            period=Period(start=date(2026, 1, 1), end=date(2027, 1, 1)),
            lines=[Line(id="gl", name="General Liability")],
            layers=[
                Layer(id="p", name="Primary", applies_to=["gl"], attach=0,
                      limit=2_000, premium=100,
                      participants=[Participant(carrier="A", share_bps=10_000)]),
                Layer(id="x", name="Excess", applies_to=["gl"], attach=2_000,
                      limit=999_998_000, premium=500_000,
                      participants=[Participant(carrier="B", share_bps=10_000)]),
            ],
        )
        _write_schematic(program, theme, tmp_path / "s.xlsx")  # must render without error
        layout, rows, _, _ = _grid(program)
        primary = next(ly for ly in layout.layers if ly.layer_id == "p")
        r0, r1 = sheet_rows(rows, primary.y0, primary.y1)
        assert r1 - r0 + 1 >= 10  # ceil(3 * 11.0 / 4.0) + 1, literal


class TestAxisLines:
    """Item 3: real Y (left, column A) and X (bottom, baseline row) axis
    lines in the theme's ink colour at medium weight."""

    def test_y_axis_left_border_spans_top_boundary_to_baseline(
        self, program, marsh, tmp_path
    ) -> None:
        path = _write_schematic(program, marsh, tmp_path / "s.xlsx")
        layout, rows, _, _ = _grid(program)
        top = max(rows.values())
        baseline_row = FIRST_GRID_ROW + top - rows[0.0] - 1
        for row in (FIRST_GRID_ROW, baseline_row):
            ref = f"{get_column_letter(AXIS_COL)}{row}"
            assert _border_hex(path, "sheet2.xml", ref, "left") == _argb(marsh.chrome.ink)

    def test_x_axis_bottom_border_spans_the_full_tower_width_at_baseline(
        self, program, marsh, tmp_path
    ) -> None:
        path = _write_schematic(program, marsh, tmp_path / "s.xlsx")
        layout, rows, _, col_of_close = _grid(program)
        xs = x_boundaries(layout)
        last_col = col_of_close[xs[-1]] - 1
        top = max(rows.values())
        baseline_row = FIRST_GRID_ROW + top - rows[0.0] - 1
        for col in (FIRST_GRID_COL, last_col):
            ref = f"{get_column_letter(col)}{baseline_row}"
            assert _border_hex(path, "sheet2.xml", ref, "bottom") == _argb(marsh.chrome.ink)

    def test_axis_lines_skip_a_retention_only_program_without_corrupting_headers(
        self, tmp_path
    ) -> None:
        """Fresh-eyes regression: a program with a retention but NO
        drawable layers has y=0.0 as its own maximum (nothing positive to
        draw), so the naive baseline_row arithmetic lands ONE ROW ABOVE
        FIRST_GRID_ROW — exactly on LINE_ROW, corrupting the line header's
        own border. There is no positive-y tower to axis in this case, so
        both axis edges must be skipped entirely rather than landing on
        whatever row the arithmetic produces."""
        theme = load_theme(None)
        program = Program(
            insured="T", program="T", placement=Placement.BOUND,
            period=Period(start=date(2026, 1, 1), end=date(2027, 1, 1)),
            lines=[Line(id="gl", name="General Liability")],
            layers=[],
            retentions=[Retention(applies_to=["gl"], type="sir", amount=250_000)],
        )
        path = _write_schematic(program, theme, tmp_path / "s.xlsx")
        ws = load_workbook(path)["T Schematic"]
        assert ws.cell(row=LINE_ROW, column=FIRST_GRID_COL).value == "General Liability"
        header_ref = f"{get_column_letter(FIRST_GRID_COL)}{LINE_ROW}"
        assert _border_hex(path, "sheet2.xml", header_ref, "bottom") is None


class TestFreezePanes:
    """Item 4: freeze at the first tower column / first grid row below
    headers, so the axis column and header rows stay visible while
    scrolling the tower."""

    def test_freeze_panes_cell_is_the_first_tower_grid_cell(
        self, program, marsh, tmp_path
    ) -> None:
        path = _write_schematic(program, marsh, tmp_path / "s.xlsx")
        ws = load_workbook(path)["Casualty Schematic"]
        assert ws.freeze_panes == f"{get_column_letter(FIRST_GRID_COL)}{FIRST_GRID_ROW}"


class TestLineSpacers:
    """Item 5: a narrow fixed-width spacer column between adjacent lines
    of cover, even where layout.py joins them flush (same group bucket,
    or both ungrouped — see layout.py's `_columns` docstring)."""

    def test_spacer_column_separates_two_joined_ungrouped_lines(
        self, tmp_path
    ) -> None:
        theme = load_theme(None)
        program = _uneven_towers_program()  # both ungrouped -> layout.py joins them
        path = _write_schematic(program, theme, tmp_path / "s.xlsx")
        ws = load_workbook(path)["T Schematic"]
        layout, _, col_of, col_of_close = _grid(program)
        tall = next(c for c in layout.columns if c.line_id == "tall")
        short = next(c for c in layout.columns if c.line_id == "short")
        assert tall.ex1 == short.ex0  # confirms layout.py joined them flush
        tall_close = col_of_close[tall.ex1]
        short_open = col_of[short.ex0]
        assert short_open == tall_close + 1  # exactly one injected slot between them
        spacer_letter = get_column_letter(tall_close)
        assert ws.column_dimensions[spacer_letter].width == pytest.approx(SPACER_UNITS)

    def test_blanket_layer_merges_straight_through_a_line_spacer(
        self, tmp_path
    ) -> None:
        """A layer that genuinely spans multiple joined lines of cover (a
        blanket primary across gl/al/el) must still merge as ONE
        continuous range — the spacer sits BETWEEN independent lines, it
        must not fracture a rect that spans across the join."""
        theme = load_theme(None)
        program = _three_column_group_program()
        path = _write_schematic(program, theme, tmp_path / "s.xlsx")
        ws = load_workbook(path)["T Schematic"]
        layout, rows, col_of, col_of_close = _grid(program)
        block = next(b for b in layout.participants)
        rect = block.rects[0]
        r0, r1 = sheet_rows(rows, rect.y0, rect.y1)
        c0, c1 = col_of[rect.x0], col_of_close[rect.x1] - 1
        ref = f"{get_column_letter(c0)}{r0}:{get_column_letter(c1)}{r1}"
        assert ref in {str(r) for r in ws.merged_cells.ranges}


class TestPrintSetup:
    """Item 6: landscape, fit-to-width one page, comfortably zoomed."""

    def test_landscape_fit_to_width_one_page(self, program, marsh, tmp_path) -> None:
        path = _write_schematic(program, marsh, tmp_path / "s.xlsx")
        ws = load_workbook(path)["Casualty Schematic"]
        assert ws.page_setup.orientation == "landscape"
        assert ws.page_setup.fitToWidth == 1
        assert ws.page_setup.fitToHeight == 0
        assert ws.sheet_view.zoomScale == 80


class TestGridlines:
    """Item 7: a faint hairline top border at every attachment boundary —
    the same boundaries the axis labels mark — but ONLY through empty grid
    cells, never through a filled participant block."""

    def test_gridline_marks_an_empty_cell_but_not_a_filled_one(
        self, tmp_path
    ) -> None:
        theme = load_theme(None)
        program = _uneven_towers_program()
        path = _write_schematic(program, theme, tmp_path / "s.xlsx")
        layout, _, col_of, _ = _grid(program)
        tall_col = col_of[next(c for c in layout.columns if c.line_id == "tall").ex0]
        short_col = col_of[next(c for c in layout.columns if c.line_id == "short").ex0]
        # $10M (tall-x's top) is the sheet's topmost grid row: "tall" is
        # filled there (tall-x), "short" (capped at $1M) is genuinely
        # blank there.
        empty_ref = f"{get_column_letter(short_col)}{FIRST_GRID_ROW}"
        filled_ref = f"{get_column_letter(tall_col)}{FIRST_GRID_ROW}"
        assert _border_hex(path, "sheet2.xml", empty_ref, "top") == _argb(theme.chrome.grid)
        assert _border_hex(path, "sheet2.xml", filled_ref, "top") != _argb(theme.chrome.grid)


class TestProvenanceFooter:
    """Item 8: a small dim provenance line below the tower, version-only
    (no git sha) so it stays git-state-independent by construction — the
    SCHEMATIC_GOLDEN_SHA and two-run byte-identity tests both depend on
    that (see tests/test_soi_xlsx.py)."""

    def test_footer_shows_the_version_only_provenance_line(
        self, program, marsh, tmp_path
    ) -> None:
        path = _write_schematic(program, marsh, tmp_path / "s.xlsx")
        ws = load_workbook(path)["Casualty Schematic"]
        values = {
            c.value for row in ws.iter_rows() for c in row if c.value is not None
        }
        assert f"towerkit {__version__} · Schedule of Insurance schematic" in values
        # git-state-independent by construction: no sha/dirty token, ever.
        assert not any(
            isinstance(v, str) and ("+dirty" in v or "unversioned" in v) for v in values
        )

    def test_footer_sits_below_the_tower_with_no_fill_or_border(
        self, program, marsh, tmp_path
    ) -> None:
        path = _write_schematic(program, marsh, tmp_path / "s.xlsx")
        ws = load_workbook(path)["Casualty Schematic"]
        layout, rows, _, _ = _grid(program)
        top = max(rows.values())
        bottom_row = FIRST_GRID_ROW + top - 1
        footer_cell = next(
            c for row in ws.iter_rows() for c in row
            if isinstance(c.value, str) and c.value.startswith("towerkit ")
        )
        assert footer_cell.row > bottom_row
        assert footer_cell.fill.patternType is None
        assert _fill_hex(path, "sheet2.xml", footer_cell.coordinate) is None


def _statutory_program() -> Program:
    return Program(
        insured="T",
        program="T",
        placement=Placement.BOUND,
        period=Period(start=date(2026, 1, 1), end=date(2027, 1, 1)),
        lines=[Line(id="wc", name="Workers Compensation")],
        layers=[Layer(
            id="wc-stat", name="Workers Compensation", applies_to=["wc"],
            attach=0, limit=0, statutory=True,
            participants=[Participant(carrier="Travelers", share_bps=10_000)],
        )],
    )


class TestStatutory:
    def test_caret_band_is_written_above_the_tower(self, marsh) -> None:
        wb = Workbook()
        wb.remove(wb.active)
        add_schematic_sheet(wb, _statutory_program(), marsh)
        ws = wb.worksheets[0]
        carets = [
            c.value for row in ws.iter_rows() for c in row
            if isinstance(c.value, str) and set(c.value) == {"^"}
        ]
        assert carets, "no caret band written"

    def test_statutory_block_has_no_top_border(self, marsh) -> None:
        wb = Workbook()
        wb.remove(wb.active)
        add_schematic_sheet(wb, _statutory_program(), marsh)
        ws = wb.worksheets[0]
        filled = [
            c for row in ws.iter_rows() for c in row
            if c.fill is not None and c.fill.fgColor is not None
            and c.border.left is not None and c.border.left.style == "thin"
        ]
        assert filled
        assert all(c.border.top is None or c.border.top.style is None for c in filled)

    def test_existing_programs_keep_their_geometry(self, program, marsh) -> None:
        """The band is added only when a statutory layer exists. A program
        without one must quantize to exactly the rows it did before.

        `max(rows.values()) == TOTAL_ROWS` alone is NOT mutation-sensitive
        here: quantize_boundaries always maps whatever the largest distinct
        boundary happens to be to TOTAL_ROWS, so a leaked extra edge past
        y=1.0 (e.g. the chevron band's own y1=1.04, if x/y_boundaries ever
        stopped gating on `layout.chevrons`) would silently re-satisfy this
        assertion while shifting every other row underneath it — exactly
        the "band leaked into geometry it should not touch" failure this
        test exists to catch. Pin the known top boundary (1.0) to
        TOTAL_ROWS explicitly so such a leak is caught here, not just by
        the golden-SHA test."""
        layout = build_layout(program)
        assert layout.chevrons == ()
        ys = y_boundaries(layout)
        assert ys[-1] == 1.0
        rows = quantize_boundaries(ys)
        assert rows[1.0] == TOTAL_ROWS
        assert max(rows.values()) == TOTAL_ROWS


# --- statutory sharing a line of cover -------------------------------------
#
# A statutory layer is drawn floor-to-top over its WHOLE column (layout.py's
# `y0, y1 = 0.0, 1.0`, off the dollar scale). A dollar-limited layer on the
# SAME line is drawn in the same column at its own y-map height, so the two
# rects genuinely overlap. Before the guard this reached openpyxl as
# `AttributeError: 'MergedCell' object attribute 'value' is read-only` from
# inside `_block` — and, when the two rects coincided exactly, as no error at
# all and a silently overpainted deliverable. validate.py already names this
# shape `statutory-line-shared`; the renderer now refuses it too, because
# `add_schematic_sheet` is a library entry point that bookkit and the MCP
# server reach without going through the CLI's validation gate.

_WC = Line(id="wc", name="Workers Compensation")
_GL = Line(id="gl", name="General Liability")


def _part_a() -> Layer:
    return Layer(
        id="wc-a", name="WC Part A", applies_to=["wc"], attach=0, limit=0,
        statutory=True,
        participants=[Participant(carrier="Travelers", share_bps=10_000)],
    )


def _part_b() -> Layer:
    return Layer(
        id="wc-b", name="Employers Liability", applies_to=["wc"], attach=0,
        limit=1_000_000,
        participants=[Participant(carrier="Travelers", share_bps=10_000)],
    )


def _gl_primary() -> Layer:
    return Layer(
        id="gl-p", name="Primary GL", applies_to=["gl"], attach=0,
        limit=2_000_000,
        participants=[Participant(carrier="Chubb", share_bps=10_000)],
    )


def _casualty(layers: list[Layer], lines: list[Line]) -> Program:
    return Program(
        insured="T", program="Casualty", placement=Placement.BOUND,
        period=Period(start=date(2026, 1, 1), end=date(2027, 1, 1)),
        lines=lines, layers=layers,
    )


def _render(program: Program, theme) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    add_schematic_sheet(wb, program, theme)
    return wb


def _drawn_cells(wb: Workbook) -> int:
    ws = wb.worksheets[0]
    return sum(
        1 for row in ws.iter_rows() for c in row
        if c.value is not None or (c.fill is not None and c.fill.patternType)
    )


class TestStatutorySharedLine:
    def test_part_a_plus_part_b_plus_gl_refuses_by_name(self, marsh) -> None:
        """The reported crash, exactly: statutory + a second layer on the
        same line + a second line of cover. The second line is what makes
        the two wc rects DIFFERENT heights (a second dollar breakpoint
        enters the global y-map), so the inner rect's anchor lands inside
        the outer rect's merge rather than on it."""
        program = _casualty([_part_a(), _part_b(), _gl_primary()], [_WC, _GL])
        with pytest.raises(SchematicOverlapError) as excinfo:
            _render(program, marsh)
        message = str(excinfo.value)
        assert "Employers Liability" in message  # the block being drawn
        assert "WC Part A" in message            # the block already there
        assert re.search(r"overlap at [A-Z]+\d+", message), message

    def test_part_a_plus_part_b_alone_refuses_too(self, marsh) -> None:
        """Without a second line the two wc rects quantize to the SAME
        range, so openpyxl never complained — `merge_cells` is idempotent
        and (r0, c0) stays the anchor — and the sheet shipped with one
        block painted straight over the other. Same invalid geometry, so
        the same refusal: a wrong client deliverable is worse than none."""
        program = _casualty([_part_a(), _part_b()], [_WC])
        with pytest.raises(SchematicOverlapError):
            _render(program, marsh)

    def test_the_refusal_matches_what_the_validator_already_says(self) -> None:
        """The renderer refuses exactly the shape validate.py names, so the
        message's `towerctl validate` pointer leads somewhere useful."""
        program = _casualty([_part_a(), _part_b(), _gl_primary()], [_WC, _GL])
        codes = {d.code for d in validate_program(program).errors}
        assert "statutory-line-shared" in codes

    @pytest.mark.parametrize(
        ("name", "layers", "lines"),
        [
            ("statutory alone", [_part_a()], [_WC]),
            ("statutory + a second line", [_part_a(), _gl_primary()], [_WC, _GL]),
            ("two layers on one line, no statutory",
             [_part_b(), _gl_primary()], [_WC, _GL]),
        ],
    )
    def test_neighbouring_shapes_still_render(self, marsh, name, layers, lines) -> None:
        """The guard must be reachable ONLY by real overlap. These three are
        the shapes one edit away from the crash; each must still produce a
        drawn tower."""
        wb = _render(_casualty(layers, lines), marsh)
        assert _drawn_cells(wb) > 0, name

    def test_a_clean_program_never_trips_the_guard(self, program, marsh) -> None:
        """`occupied` now carries a label per cell and every `_block` call
        checks it; the ordinary multi-line, multi-layer program with
        retentions must be untouched by that."""
        assert _drawn_cells(_render(program, marsh)) > 0

    def test_overlap_below_a_free_anchor_is_caught_too(self, marsh) -> None:
        """The guard scans the WHOLE range, not just the anchor cell — and
        this is the shape that needs it. A blanket primary spanning both
        lines is drawn FIRST (bottom half of both columns), then the
        statutory bar's own rect runs the full height of the wc column:
        its anchor (top-left) is free, and the collision only starts
        halfway down.

        Unguarded, this one does not even raise. `ws.cell(r0, c0)` finds a
        real Cell, so openpyxl's read-only MergedCell never comes up — the
        render "succeeds" and writes a workbook with two OVERLAPPING merged
        ranges, which Excel opens as damaged or silently repairs. That is
        strictly worse than the reported crash, so an anchor-only check
        would not be enough."""
        blanket = Layer(
            id="blk", name="Blanket Primary", applies_to=["wc", "gl"],
            attach=0, limit=1_000_000,
            participants=[Participant(carrier="Chubb", share_bps=10_000)],
        )
        excess = Layer(
            id="gl-x", name="GL Excess", applies_to=["gl"],
            attach=1_000_000, limit=1_000_000,
            participants=[Participant(carrier="AIG", share_bps=10_000)],
        )
        program = _casualty([blanket, excess, _part_a()], [_WC, _GL])
        layout = build_layout(program)
        rows = quantize_boundaries(
            y_boundaries(layout), label_spans=_label_spans(layout, program)
        )
        col_of, col_of_close = line_column_slots(layout, x_boundaries(layout))
        drawn: dict[tuple[int, int], str] = {}
        anchors: dict[str, tuple[int, int]] = {}
        for block in layout.participants:
            for rect in block.rects:
                r0, r1 = sheet_rows(rows, rect.y0, rect.y1)
                c0, c1 = col_of[rect.x0], col_of_close[rect.x1] - 1
                anchors.setdefault(block.layer_id, (r0, c0))
                for r in range(r0, r1 + 1):
                    for c in range(c0, c1 + 1):
                        drawn.setdefault((r, c), block.layer_id)
        # the premise: the statutory bar's anchor is NOT the colliding cell
        assert drawn[anchors["wc-a"]] == "wc-a"
        with pytest.raises(SchematicOverlapError) as excinfo:
            _render(program, marsh)
        assert "Blanket Primary" in str(excinfo.value)
