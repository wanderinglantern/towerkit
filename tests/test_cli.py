"""End-to-end CLI: named subcommands, never positional sys.argv."""

from pathlib import Path

from towerkit.cli import main

REPO = Path(__file__).parent.parent
OLD = REPO / "programs" / "atomic-2026.json"
NEW = REPO / "programs" / "atomic-2027.json"


class TestRender:
    def test_render_writes_requested_formats(self, tmp_path, capsys) -> None:
        code = main(
            [
                "render", str(OLD),
                "--theme", str(REPO / "themes" / "marsh.json"),
                "--out", str(tmp_path),
                "--format", "svg,png",
            ]
        )
        assert code == 0
        assert (tmp_path / "atomic-2026.svg").exists()
        assert (tmp_path / "atomic-2026.png").exists()
        out = capsys.readouterr().out
        assert "atomic-2026.svg" in out

    def test_render_refuses_invalid_program(self, tmp_path) -> None:
        bad = tmp_path / "bad.json"
        bad.write_text(OLD.read_text().replace('"attach": 27000000', '"attach": 30000000'))
        code = main(["render", str(bad), "--out", str(tmp_path)])
        assert code == 1
        assert not (tmp_path / "bad.svg").exists()


class TestCompare:
    def test_compare_writes_comparison(self, tmp_path) -> None:
        code = main(
            ["compare", str(OLD), str(NEW), "--out", str(tmp_path), "--format", "svg"]
        )
        assert code == 0
        assert (tmp_path / "atomic-2026-vs-atomic-2027.svg").exists()


def test_soi_exports_workbook(tmp_path) -> None:
    sample = Path(__file__).parent.parent / "programs" / "atomic-2026.json"
    out = tmp_path / "soi.xlsx"
    assert main(["soi", str(sample), "-o", str(out)]) == 0
    assert out.exists() and out.stat().st_size > 0


def test_soi_default_filename_from_insured(tmp_path, monkeypatch) -> None:
    from towerkit.model import load_program
    from towerkit.soi import default_filename

    sample = Path(__file__).parent.parent / "programs" / "atomic-2026.json"
    monkeypatch.chdir(tmp_path)
    assert main(["soi", str(sample)]) == 0
    assert (tmp_path / default_filename(load_program(sample))).exists()


class TestImport:
    def test_template_then_import(self, tmp_path, capsys) -> None:
        import json

        template = tmp_path / "t.xlsx"
        assert main(["template", str(template)]) == 0
        out = tmp_path / "prog.json"
        code = main(
            [
                "import", str(template), "-o", str(out),
                "--insured", "Example Co", "--program", "Property",
            ]
        )
        assert code == 0
        data = json.loads(out.read_text())
        assert data["insured"] == "Example Co"
        assert len(data["layers"]) == 2

    def test_import_paste_without_period_refuses(self, tmp_path) -> None:
        paste = tmp_path / "tower.txt"
        paste.write_text("Primary 10M — Chubb 100% — 250,000\n5M xs 10M — Sompo\n")
        out = tmp_path / "prog.json"
        code = main(
            [
                "import", str(paste), "-o", str(out),
                "--insured", "Example Co", "--program", "Casualty",
            ]
        )
        assert code == 1  # no period in pasted text → error, nothing written
        assert not out.exists()

    def test_import_paste_with_dates(self, tmp_path) -> None:
        import json

        paste = tmp_path / "tower.txt"
        paste.write_text("Primary 10M — Chubb 100%\n")
        out = tmp_path / "prog.json"
        code = main(
            [
                "import", str(paste), "-o", str(out), "--insured", "Example Co",
                "--program", "Casualty",
                "--inception", "10/1/2026", "--expiry", "10/1/2027",
            ]
        )
        assert code == 0
        assert json.loads(out.read_text())["period"]["start"] == "2026-10-01"

    def test_import_csv_unknown_column_exits_nonzero(self, tmp_path, capsys) -> None:
        bad = tmp_path / "bad.csv"
        bad.write_text("layer,limit,wibble\nPrimary,10M,x\n")
        code = main(
            [
                "import", str(bad), "-o", str(tmp_path / "x.json"),
                "--insured", "A", "--program", "P",
            ]
        )
        assert code == 1
        assert "wibble" in capsys.readouterr().out

    def test_import_refuses_existing_output(self, tmp_path, capsys) -> None:
        out = tmp_path / "existing.json"
        out.write_text("{}")  # a program file is the source of truth
        paste = tmp_path / "tower.txt"
        paste.write_text("Primary 10M — Chubb 100%\n")
        code = main(
            [
                "import", str(paste), "-o", str(out), "--insured", "A",
                "--program", "P", "--inception", "2026-10-01", "--expiry", "2027-10-01",
            ]
        )
        assert code == 1
        assert out.read_text() == "{}"  # untouched

    def test_import_bad_rows_exits_nonzero(self, tmp_path, capsys) -> None:
        bad = tmp_path / "bad.csv"
        bad.write_text(
            "layer,limit,attachment,carrier,inception,expiry\n"
            "Primary,banana,0,Chubb,2026-10-01,2027-10-01\n"
        )
        code = main(
            [
                "import", str(bad), "-o", str(tmp_path / "x.json"),
                "--insured", "A", "--program", "P",
            ]
        )
        assert code == 1
        assert "banana" in capsys.readouterr().out


class TestImportDiagnostics:
    @staticmethod
    def _template_with(tmp_path, column: str, value: str):
        """The shipped template, with one cell of row 2 overwritten."""
        from openpyxl import load_workbook

        from towerkit.ingest_template import write_template

        source = write_template(tmp_path / "sched.xlsx")
        wb = load_workbook(source)
        ws = wb.worksheets[0]
        headers = [str(c.value or "").strip().lower() for c in ws[1]]
        ws.cell(row=2, column=headers.index(column) + 1, value=value)
        wb.save(source)
        return source

    def test_warnings_are_printed_and_the_file_is_still_written(
        self, tmp_path, capsys
    ) -> None:
        """A partly-placed tower is a legitimate import — it must be
        written, and it must not go out silently."""
        source = self._template_with(tmp_path, "share", "60%")
        out = tmp_path / "imported.json"

        code = main(
            [
                "import", str(source), "-o", str(out),
                "--insured", "Example Co", "--program", "Property",
            ]
        )

        captured = capsys.readouterr()
        assert code == 0
        assert out.exists()
        assert "placed" in captured.out, captured.out

    def test_errors_stop_the_import_and_exit_non_zero(self, tmp_path) -> None:
        # A bad "share" cell, not a bad "limit": a bad limit drops the
        # whole layer during row parsing, which cascades into a
        # validate_program line-base failure — to_program() raises
        # ProgramInvalidError and lands on the same (already-covered)
        # `program is None` branch as a missing policy period. A bad
        # share only drops that row's participant, leaving an unplaced
        # layer that still validates (unplaced is a warning, not an
        # error), so to_program() SUCCEEDS while draft.diagnostics.errors
        # stays non-empty from the row-level parse error. That is the
        # only input that reaches `_cmd_import`'s
        # `program is None or draft.diagnostics.errors` guard through its
        # second half instead of its first — verified by instrumenting a
        # run, not by reading the parser.
        source = self._template_with(tmp_path, "share", "banana%")
        out = tmp_path / "imported.json"

        code = main(
            [
                "import", str(source), "-o", str(out),
                "--insured", "Example Co", "--program", "Property",
            ]
        )

        assert code == 1
        assert not out.exists(), "a schedule with errors must not be written"


class TestParser:
    def test_no_command_shows_help(self, capsys) -> None:
        assert main([]) == 2

    def test_version(self, capsys) -> None:
        try:
            main(["--version"])
        except SystemExit as exc:
            assert exc.code == 0
        assert "towerctl" in capsys.readouterr().out


def test_soi_schematic_flag_adds_second_sheet(tmp_path) -> None:
    from openpyxl import load_workbook

    out = tmp_path / "soi.xlsx"
    assert main(["soi", str(OLD), "-o", str(out), "--schematic"]) == 0
    assert len(load_workbook(out).sheetnames) == 2


def test_soi_without_flag_stays_single_sheet(tmp_path) -> None:
    from openpyxl import load_workbook

    out = tmp_path / "soi.xlsx"
    assert main(["soi", str(OLD), "-o", str(out)]) == 0
    assert len(load_workbook(out).sheetnames) == 1


def test_soi_honors_stored_render_setting(tmp_path) -> None:
    import json

    from openpyxl import load_workbook

    data = json.loads(OLD.read_text())
    data.setdefault("render", {})["soiSchematic"] = True
    sample = tmp_path / "stored.json"
    sample.write_text(json.dumps(data))
    out = tmp_path / "soi.xlsx"
    assert main(["soi", str(sample), "-o", str(out)]) == 0
    assert len(load_workbook(out).sheetnames) == 2
