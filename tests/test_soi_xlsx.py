"""SOI workbook writer: styling mirrors the sample; output is byte-identical."""

import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

import pytest
from openpyxl import load_workbook
from test_soi import make_program, unbound_with_premium

from towerkit.render.soi_xlsx import render_soi_sheet, write_soi, write_soi_workbook
from towerkit.soi import build_soi, sheet_title
from towerkit.theme import load_theme

_SML_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"


@pytest.fixture()
def program():
    return make_program()


@pytest.fixture()
def theme():
    return load_theme(None)


def _write(program, theme, path: Path, **kw) -> Path:
    return write_soi(
        build_soi(program), title=sheet_title(program), theme=theme, out_path=path, **kw
    )


def _core_xml(xlsx_path: Path) -> str:
    with zipfile.ZipFile(xlsx_path) as z:
        return z.read("docProps/core.xml").decode()


def _row2_fill_hexes(xlsx_path: Path) -> list[str | None]:
    """Per-column fill (as ARGB hex, or None if unfilled) for sheet row 2,
    read straight from the saved XML.

    openpyxl's reader doesn't reliably expose style on the non-anchor cells
    of a merged range: reloading the same file and inspecting e.g. ws["D2"]
    (a MergedCell) always reports patternType=None regardless of what's on
    disk, even though the underlying <c> element carries a real fillId. So
    this reads xl/worksheets/sheet1.xml + xl/styles.xml directly rather than
    going through load_workbook — see the Task 5 fix report.
    """
    with zipfile.ZipFile(xlsx_path) as z:
        styles = ET.fromstring(z.read("xl/styles.xml"))
        sheet = ET.fromstring(z.read("xl/worksheets/sheet1.xml"))

    fills = []
    for fill in styles.find(f"{_SML_NS}fills"):
        pattern = fill.find(f"{_SML_NS}patternFill")
        fg = pattern.find(f"{_SML_NS}fgColor") if pattern is not None else None
        fills.append(fg.get("rgb") if fg is not None else None)
    fill_ids = [int(xf.get("fillId", "0")) for xf in styles.find(f"{_SML_NS}cellXfs")]

    row2 = next(r for r in sheet.iter(f"{_SML_NS}row") if r.get("r") == "2")
    hexes = []
    for cell in row2.findall(f"{_SML_NS}c"):
        s = int(cell.get("s", "0"))
        hexes.append(fills[fill_ids[s]])
    return hexes


class TestDeterminism:
    def test_two_writes_are_byte_identical(self, program, theme, tmp_path) -> None:
        a = _write(program, theme, tmp_path / "a.xlsx")
        b = _write(program, theme, tmp_path / "b.xlsx")
        assert a.read_bytes() == b.read_bytes()

    def test_modified_timestamp_is_pinned_not_wall_clock(self, program, theme, tmp_path) -> None:
        # A same-second write can pass the byte-identity test by coincidence
        # even if <dcterms:modified> isn't actually neutralized. Assert the
        # pinned value directly so the test can't pass by timing luck.
        path = _write(program, theme, tmp_path / "s.xlsx")
        assert "<dcterms:modified" in _core_xml(path)
        assert "1980-01-01T00:00:00Z</dcterms:modified>" in _core_xml(path)


class TestContent:
    def test_headers_and_title(self, program, theme, tmp_path) -> None:
        ws = load_workbook(_write(program, theme, tmp_path / "s.xlsx")).active
        assert ws.title == "Casualty SOI - 26-27"
        assert [c.value for c in ws[1]] == [
            "Insured", "Line of Coverage", "Status", "Carrier", "Policy Number",
            "Effective Date", "Expiration Date", "Limits",
            "Deductible / SIR / Retention", "Premium",
        ]

    def test_header_style_mirrors_sample(self, program, theme, tmp_path) -> None:
        ws = load_workbook(_write(program, theme, tmp_path / "s.xlsx")).active
        cell = ws["A1"]
        assert cell.fill.fgColor.rgb == "FF003865"
        assert cell.font.bold and cell.font.name == "Noto Sans"
        assert cell.alignment.horizontal == "center" and cell.alignment.wrap_text
        assert ws.freeze_panes == "A2"

    def test_section_band_row_is_a_full_width_label(self, program, theme, tmp_path) -> None:
        """The mingled roll-up that used to sit in the band's last column is
        gone: money now lives on the two subtotal lines beneath the rows, so
        the reader is never shown bound and unbound added together."""
        ws = load_workbook(_write(program, theme, tmp_path / "s.xlsx")).active
        assert ws["A2"].value == "Casualty"
        assert "A2:J2" in {str(r) for r in ws.merged_cells.ranges}
        assert ws["J2"].value is None

    def test_section_band_fill_covers_every_column(self, program, theme, tmp_path) -> None:
        # Merged A2:H2 plus the I2 roll-up should all carry the accent fill —
        # one solid band, not a navy patch at each end with white in between.
        path = _write(program, theme, tmp_path / "s.xlsx")
        hexes = _row2_fill_hexes(path)
        assert len(hexes) == 10  # A..J
        assert hexes == ["FF003865"] * 10

    def test_zebra_restarts_per_section(self, program, theme, tmp_path) -> None:
        ws = load_workbook(_write(program, theme, tmp_path / "s.xlsx")).active
        # Casualty: band r2, body r3-r6, subtotals r7-r8; next body starts r9.
        assert ws["A3"].fill.patternType is None          # first body row: white
        assert ws["A4"].fill.fgColor.rgb == "FFF7F3EE"    # second: banded
        assert ws["A9"].fill.patternType is None          # new section restarts white

    def test_unlabeled_section_has_no_band_row(self, program, theme, tmp_path) -> None:
        ws = load_workbook(_write(program, theme, tmp_path / "s.xlsx")).active
        assert ws["B9"].value == "Property — Primary"  # body row, not a section label

    def test_dates_are_real_dates(self, program, theme, tmp_path) -> None:
        import datetime

        ws = load_workbook(_write(program, theme, tmp_path / "s.xlsx")).active
        cell = ws["F3"]
        assert cell.value == datetime.datetime(2026, 2, 1)
        assert cell.number_format == "mm/dd/yyyy"

    def test_no_premiums_drops_the_column(self, program, theme, tmp_path) -> None:
        ws = load_workbook(
            _write(program, theme, tmp_path / "s.xlsx", show_premiums=False)
        ).active
        assert [c.value for c in ws[1]][-1] == "Deductible / SIR / Retention"
        assert ws.max_column == 9
        assert "A2:I2" in {str(r) for r in ws.merged_cells.ranges}  # full-width band
        # no premium column, so no subtotal lines: the next section's body
        # follows the Casualty rows immediately
        assert ws["B7"].value == "Property — Primary"


class TestStatusAndSubtotals:
    """C1/C2: the schedule says, per row, whether the cover exists — and the
    premium the reader adds up is only the premium for cover that does."""

    def test_status_renders_per_row(self, program, theme, tmp_path) -> None:
        ws = load_workbook(_write(program, theme, tmp_path / "s.xlsx")).active
        assert [ws[f"C{r}"].value for r in (3, 4, 5, 6)] == [
            "Bound", "Bound", "Bound", "To be placed",
        ]

    def test_subtotal_lines_sit_under_the_section_rows(self, program, theme, tmp_path):
        ws = load_workbook(_write(program, theme, tmp_path / "s.xlsx")).active
        assert ws["A7"].value == "Bound cover — premium subtotal"
        assert ws["A8"].value == "Unbound cover — premium subtotal"
        assert "A7:I7" in {str(r) for r in ws.merged_cells.ranges}
        assert ws["J7"].number_format == '"$"#,##0.00'

    def test_an_unbound_premium_is_never_inside_the_bound_subtotal(
        self, theme, tmp_path
    ) -> None:
        """THE load-bearing assertion: the "To be placed" layer carries
        $25,000 and it must not be inside the line labelled Bound."""
        ws = load_workbook(
            _write(unbound_with_premium(), theme, tmp_path / "s.xlsx")
        ).active
        assert ws["C6"].value == "To be placed" and ws["J6"].value == 25_000
        assert ws["J7"].value == 100_000   # bound cover only
        assert ws["J8"].value == 25_000    # unbound, stated separately

    def test_a_wholly_unstated_subtotal_prints_an_em_dash_not_zero(
        self, program, theme, tmp_path
    ) -> None:
        """C, fix round 1: Casualty's only unbound row carries NO premium, so
        the cell used to read $0.00 — the rendering the body refuses one row
        higher up, because $0.00 reads as free cover. Under a visible "To be
        placed" row it said the unplaced cover was free."""
        ws = load_workbook(_write(program, theme, tmp_path / "s.xlsx")).active
        assert ws["C6"].value == "To be placed" and ws["J6"].value is None
        assert ws["A8"].value == "Unbound cover — premium subtotal"
        assert ws["J8"].value == "—"

    def test_a_section_with_nothing_unbound_also_states_nothing(
        self, program, theme, tmp_path
    ) -> None:
        ws = load_workbook(_write(program, theme, tmp_path / "s.xlsx")).active
        assert ws["B9"].value == "Property — Primary"   # the section's only row
        assert ws["J10"].value == 80_000                # bound, a real figure
        assert ws["J11"].value == "—"                   # nothing unbound to state

    def test_an_unlabelled_section_still_gets_its_subtotal_lines(
        self, program, theme, tmp_path
    ) -> None:
        """E, fix round 1: subtotals under a section with no band row was a
        judgement call held ONLY by the golden hash. Pinned by name now, so a
        regeneration cannot drop it silently."""
        ws = load_workbook(_write(program, theme, tmp_path / "s.xlsx")).active
        assert ws["B9"].value == "Property — Primary"   # body row, no band above
        assert ws["A10"].value == "Bound cover — premium subtotal"
        assert ws["A11"].value == "Unbound cover — premium subtotal"
        assert ws["A12"].value == "Program-wide"        # next section's band

    def test_no_premiums_drops_the_subtotal_lines(self, program, theme, tmp_path):
        ws = load_workbook(
            _write(program, theme, tmp_path / "s.xlsx", show_premiums=False)
        ).active
        assert "subtotal" not in str([c.value for c in ws["A"]])

    def test_a_zero_premium_reads_included_not_free(self, theme, tmp_path) -> None:
        p = make_program()
        next(x for x in p.layers if x.id == "al-primary").premium = 0
        ws = load_workbook(_write(p, theme, tmp_path / "s.xlsx")).active
        assert ws["J6"].value == "Included"


class TestRowHeightColumnIndices:
    """_LIMITS_IX / _RETENTION_IX index into the RENDERED row tuple, and they
    shifted 6,7 -> 7,8 when the Status column was inserted. Nothing pinned
    them: no fixture row has limits over 100 characters or retention over 34,
    so every row lands on the 36.0 floor whatever the constants say, and a
    future column insert would shift them again with nothing to notice.

    These two rows are long enough that reading the wrong column collapses the
    height to a different number (fix round 1). They are deliberately NOT in
    the golden fixture: a fixture row exists to be read by twenty other
    assertions, and lengthening one to close a hole here would move the golden
    for a test-only reason."""

    def test_long_limits_prose_drives_the_row_height(self, theme, tmp_path) -> None:
        p = make_program()
        p.layers[0].limits_detail = "X" * 250   # 3 wrapped lines at width 100
        ws = load_workbook(_write(p, theme, tmp_path / "s.xlsx")).active
        assert ws["H3"].value == "X" * 250      # column H IS the limits column
        assert ws.row_dimensions[3].height == 54.0

    def test_long_retention_prose_drives_the_row_height(self, theme, tmp_path) -> None:
        p = make_program()
        p.layers[0].retention_detail = "Y" * 105   # 4 wrapped lines at width 34
        ws = load_workbook(_write(p, theme, tmp_path / "s.xlsx")).active
        assert ws["I3"].value == "Y" * 105         # column I IS the retention column
        assert ws.row_dimensions[3].height == 72.0


# Refactor guard: extracting the generic table writer (render/table_xlsx.py)
# must not change SOI output. docProps/core.xml embeds provenance() — the
# CURRENT git sha and dirty marker — so RAW file bytes change with every
# commit; the guard therefore hashes every zip entry EXCEPT core.xml.
# Regenerate GOLDEN_SHA only on a deliberate style/content change or an
# openpyxl bump — never to make a refactor pass.
GOLDEN_SHA = "39e0fa05c54056bb3f887473d37fd48b167764acd55a0ce9e2abeeed82a13615"
# Regenerated 2026-08-18 (soi-status-and-statutory, fix round 1): the unbound
# subtotal of a section whose unbound rows state no premium prints an em dash
# instead of $0.00 — Casualty's only unbound row has premium=None, so the sheet
# said "Unbound cover — premium subtotal $0.00" beneath a visible "To be
# placed" row, which is the free-cover reading premium_value refuses one row
# higher up. Three cells move (the unbound line of all three sections) and
# nothing else does. Asserted by name in TestStatusAndSubtotals above.
#
# Regenerated 2026-08-18 (soi-status-and-statutory): a deliberate CONTENT
# change, not a refactor. The sheet gained a Status column (C1), the section
# band lost its mingled bound+unbound roll-up in favour of two labelled
# subtotal lines beneath the rows (C1/C2), and a shared retention is stated
# once instead of on every primary it touches (C10). NOT the statutory phrase
# and NOT the "Included" zero premium: this fixture has no statutory layer and
# no zero premium, so neither contributes a byte to the hash — they are pinned
# by their own named tests instead (comment corrected in fix round 1). Every
# reason listed is asserted by name above; this hash only pins nothing ELSE
# moved.


def _content_hash(xlsx_path: Path) -> str:
    import hashlib

    digest = hashlib.sha256()
    with zipfile.ZipFile(xlsx_path) as z:
        for name in sorted(z.namelist()):
            if name == "docProps/core.xml":
                continue
            digest.update(name.encode())
            digest.update(z.read(name))
    return digest.hexdigest()


def test_refactor_golden_content(program, theme, tmp_path):
    path = _write(program, theme, tmp_path / "golden.xlsx")
    assert _content_hash(path) == GOLDEN_SHA


def test_render_soi_sheet_into_an_open_workbook(program, theme, tmp_path):
    """The bookkit composition shape: a table sheet plus an SOI sheet in one
    workbook, one finalize."""
    from openpyxl import Workbook

    from towerkit.render.table_xlsx import finalize_workbook, sanitize_sheet_title
    from towerkit.soi import build_soi

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Open Items"  # stand-in for bookkit's table sheet
    soi_ws = wb.create_sheet(sanitize_sheet_title("Schedule of Insurance"))
    render_soi_sheet(soi_ws, build_soi(program), theme=theme)
    path = finalize_workbook(wb, tmp_path / "multi.xlsx")
    loaded = load_workbook(path)
    assert loaded.sheetnames == ["Open Items", "Schedule of Insurance"]
    sheet = loaded["Schedule of Insurance"]
    assert [c.value for c in sheet[1]][:3] == ["Insured", "Line of Coverage", "Status"]
    assert sheet["A2"].value == "Casualty"          # section band renders
    assert sheet.freeze_panes == "A2"
    assert sheet.row_dimensions[3].height >= 36.0   # SOI row-height heuristic applied


class TestWorkbookOrchestration:
    def test_off_is_byte_identical_to_write_soi(self, program, theme, tmp_path):
        legacy = _write(program, theme, tmp_path / "legacy.xlsx")
        combined = write_soi_workbook(
            program, theme=theme, out_path=tmp_path / "combined.xlsx",
            include_schematic=False,
        )
        assert combined.read_bytes() == legacy.read_bytes()

    def test_off_still_matches_the_golden(self, program, theme, tmp_path):
        path = write_soi_workbook(program, theme=theme, out_path=tmp_path / "g.xlsx")
        assert _content_hash(path) == GOLDEN_SHA

    def test_on_appends_the_schematic_sheet(self, program, theme, tmp_path):
        path = write_soi_workbook(
            program, theme=theme, out_path=tmp_path / "s.xlsx", include_schematic=True
        )
        assert load_workbook(path).sheetnames == [
            "Casualty SOI - 26-27", "Casualty Schematic",
        ]

    def test_on_two_runs_byte_identical(self, program, theme, tmp_path):
        a = write_soi_workbook(
            program, theme=theme, out_path=tmp_path / "a.xlsx", include_schematic=True
        )
        b = write_soi_workbook(
            program, theme=theme, out_path=tmp_path / "b.xlsx", include_schematic=True
        )
        assert a.read_bytes() == b.read_bytes()


# Content golden WITH the schematic sheet (same core.xml-exclusion mechanism
# as GOLDEN_SHA above; same regeneration rule — deliberate change or
# openpyxl bump only, never to make a refactor pass).
SCHEMATIC_GOLDEN_SHA = "deac2226cc645b75e16fd411a3a453b42880d707582976495c1c086a4f1c1668"
# Regenerated 2026-08-18 (soi-status-and-statutory, fix round 1): the SCHEMATIC
# sheet is still unchanged; this moves for exactly the reason GOLDEN_SHA moves
# in the same round (the em-dash unbound subtotal) and for no others.
# Regenerated 2026-08-18 (soi-status-and-statutory): the SCHEMATIC sheet is
# unchanged — this workbook contains the SOI sheet too, so it moves for
# exactly the reasons GOLDEN_SHA above moves and for no others.
# Regenerated 2026-08-13 (Task 11, polish round 3, feat/schematic-polish3):
# narrower canvas (CANVAS_WIDTH_UNITS 300->240), label-aware row floor,
# real axis lines, freeze panes, line-of-cover spacer columns, landscape
# print setup, faint attachment gridlines, and a version-only provenance
# footer. Deliberate, one-time regeneration — see that commit's body.


def test_schematic_golden_content(program, theme, tmp_path):
    path = write_soi_workbook(
        program, theme=theme, out_path=tmp_path / "sg.xlsx", include_schematic=True
    )
    assert _content_hash(path) == SCHEMATIC_GOLDEN_SHA


# --- CRITICAL 1: a sublimit must not disappear behind prose limits ----------
#
# Deliberately NOT the golden fixture, for the reason TestRowHeightColumnIndices
# gives: a fixture row exists to be read by twenty other assertions, and this
# defect needs a shape (prose on the layer that carries the line's cover, a
# sublimit on the line) that the golden fixture does not have.
#
# The assertions are on RENDERED CELLS, not on limits_text's return value. The
# bug is about what reaches the reader of the workbook, and the reader reads
# cells.

_PROSE = (
    "$100,000,000 per occurrence and in the annual aggregate for the perils "
    "of fire, lightning, windstorm, hail and all other perils not otherwise "
    "excluded"
)


def _prose_limits_program():
    """The reproduction: a bound property primary stating its limit in words,
    a to-be-placed excess above it, and two sublimits on the line."""
    from towerkit.model import Program

    return Program.model_validate({
        "insured": "Harbor Point Holdings, LLC",
        "program": "Property",
        "placement": "bound",
        "period": {"start": "2026-01-01", "end": "2027-01-01"},
        "lines": [{"id": "prop", "name": "Property"}],
        "layers": [
            {
                "id": "prop-primary", "name": "Primary", "appliesTo": ["prop"],
                "attach": 0, "limit": 100_000_000, "premium": 500_000,
                "limitsDetail": _PROSE,
                "participants": [
                    {"carrier": "FM Global", "share_bps": 6_000},
                    {"carrier": "Chubb", "share_bps": 4_000},
                ],
            },
            {
                "id": "prop-x1", "name": "1st Excess", "appliesTo": ["prop"],
                "attach": 100_000_000, "limit": 150_000_000,
            },
        ],
        "sublimits": [
            {"name": "Flood", "amount": 25_000_000, "appliesTo": ["prop"]},
            {"name": "Earthquake", "amount": 10_000_000, "appliesTo": ["prop"]},
        ],
    })


def _limits_cells(ws) -> dict[str, str]:
    """{coverage cell -> limits cell} for every body row on the sheet."""
    out = {}
    for row in ws.iter_rows(min_row=2):
        coverage, limits = row[1].value, row[7].value
        if coverage and limits:
            out[str(coverage)] = str(limits)
    return out


class TestSublimitsSurviveProseLimits:
    def test_the_prose_row_still_states_the_sublimits(self, theme, tmp_path) -> None:
        """H18 in the reproduction: the in-force primary printed the prose and
        NOTHING about flood, so the schedule told the reader flood was covered
        to $100,000,000 when it is covered to $25,000,000."""
        p = _prose_limits_program()
        ws = load_workbook(_write(p, theme, tmp_path / "s.xlsx")).active
        cell = _limits_cells(ws)["Property — Primary"]
        assert "Sublimit: Flood $25,000,000" in cell
        assert "Sublimit: Earthquake $10,000,000" in cell

    def test_the_prose_is_still_first_and_still_verbatim(self, theme, tmp_path):
        """Prose winning is the shipped contract and it stays: the sublimit
        tail is appended to it, never allowed to displace or reword it."""
        p = _prose_limits_program()
        ws = load_workbook(_write(p, theme, tmp_path / "s.xlsx")).active
        cell = _limits_cells(ws)["Property — Primary"]
        assert cell == (
            f"{_PROSE}; Sublimit: Flood $25,000,000; "
            "Sublimit: Earthquake $10,000,000"
        )

    def test_the_two_rows_no_longer_contradict_each_other(self, theme, tmp_path):
        """The excess row printed the sublimits all along. A schedule whose
        unplaced layer names a cap its bound layer does not is telling the
        reader two different things about one line."""
        p = _prose_limits_program()
        ws = load_workbook(_write(p, theme, tmp_path / "s.xlsx")).active
        cells = _limits_cells(ws)
        for coverage in ("Property — Primary", "Property — 1st Excess"):
            assert "Sublimit: Flood $25,000,000" in cells[coverage], coverage

    def test_prose_on_every_layer_no_longer_loses_the_sublimit_entirely(
        self, theme, tmp_path
    ) -> None:
        """The worst case: with prose on BOTH layers there was no other layer
        on the line for the tail to re-attach itself to, so the sublimit left
        the workbook without a trace."""
        p = _prose_limits_program()
        p.layers[1].limits_detail = "$150,000,000 excess of the primary layer"
        ws = load_workbook(_write(p, theme, tmp_path / "s.xlsx")).active
        sheet_text = "\n".join(_limits_cells(ws).values())
        assert sheet_text.count("Sublimit: Flood $25,000,000") == 2
        assert sheet_text.count("Sublimit: Earthquake $10,000,000") == 2

    def test_statutory_prose_keeps_its_sublimit_too(self, theme, tmp_path) -> None:
        """The statutory escape hatch is prose by another name; it must not
        become a second way to lose the tail."""
        p = _prose_limits_program()
        p.layers[0].limits_detail = None
        p.layers[0].statutory = True
        p.layers[0].limit = 0
        ws = load_workbook(_write(p, theme, tmp_path / "s.xlsx")).active
        cell = _limits_cells(ws)["Property — Primary"]
        assert cell.startswith("Statutory - State Limits")
        assert "Sublimit: Flood $25,000,000" in cell
