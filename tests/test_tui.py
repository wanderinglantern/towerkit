"""TUI tests via Textual's run_test pilot.

The headline gate: a valid program can be created from scratch without
touching a text editor — and saving is canonical, blocked-but-not-silent on
errors, with working undo/redo and over-sign protection.
"""

from __future__ import annotations

import errno
import shutil
from pathlib import Path

import pytest
from textual.widgets import Button, Input

from towerkit.model import dumps_program, load_program
from towerkit.money import BPS_SCALE
from towerkit.tui.app import TowerkitApp
from towerkit.tui.screens.browser import ProgramBrowser, _bump_stem
from towerkit.tui.screens.editor import EditorScreen
from towerkit.tui.session import EditSession, suggested_attach
from towerkit.tui.widgets.inputs import CarrierSuggester, parse_share_pct

REPO = Path(__file__).parent.parent
SAMPLE = REPO / "programs" / "atomic-2026.json"


def _filled_template(tmp_path: Path) -> Path:
    """A ready-to-import workbook. `write_template` (see
    towerkit.ingest_template) already writes the worked example rows that
    the CLI round-trip test (tests/test_cli.py::TestTemplate::
    test_template_then_import) imports as-is — no manual openpyxl filling
    step is needed on top of it."""
    from towerkit.ingest_template import write_template

    return write_template(tmp_path / "filled.xlsx")


@pytest.fixture()
def sample_copy(tmp_path):
    target = tmp_path / "programs"
    target.mkdir()
    shutil.copy(SAMPLE, target / "atomic-2026.json")
    return target / "atomic-2026.json"


class TestSession:
    def test_undo_redo_round_trip(self) -> None:
        session = EditSession.open(SAMPLE)
        original = dumps_program(session.program)
        session.mutate(lambda p: setattr(p, "insured", "Changed Co"))
        assert session.program.insured == "Changed Co"
        assert session.undo()
        assert dumps_program(session.program) == original
        assert session.redo()
        assert session.program.insured == "Changed Co"

    def test_noop_mutation_costs_no_undo_step(self) -> None:
        session = EditSession.open(SAMPLE)
        session.mutate(lambda p: None)
        assert not session.undo()

    def test_save_with_no_edits_is_zero_diff(self, sample_copy) -> None:
        before = sample_copy.read_bytes()
        session = EditSession.open(sample_copy)
        session.save()
        assert sample_copy.read_bytes() == before

    def test_dirty_tracking(self, sample_copy) -> None:
        session = EditSession.open(sample_copy)
        assert not session.dirty
        session.mutate(lambda p: setattr(p, "insured", "X Co"))
        assert session.dirty
        session.save()
        assert not session.dirty

    def test_suggested_attach_is_top_of_stack(self) -> None:
        program = load_program(SAMPLE)
        # gl/al/el stack tops out at $202M
        assert suggested_attach(program, ["gl"]) == 202_000_000
        # pl tower tops out at $25M
        assert suggested_attach(program, ["pl"]) == 25_000_000
        assert suggested_attach(program, ["nope"]) == 0

    def test_add_layer_defaults_contiguous(self) -> None:
        session = EditSession.open(SAMPLE)
        layer = session.add_layer(["pl"])
        assert layer.attach == 25_000_000  # contiguity by construction

    def test_save_refuses_when_the_file_moved_underneath(self, sample_copy) -> None:
        from towerkit.tui.session import StaleFileError

        session = EditSession.open(sample_copy)
        session.mutate(lambda p: setattr(p, "insured", "Mine"))
        sample_copy.write_text(sample_copy.read_text("utf-8") + " ", encoding="utf-8")
        with pytest.raises(StaleFileError):
            session.save()

    def test_forced_save_overwrites_and_re_arms_the_guard(self, sample_copy) -> None:
        session = EditSession.open(sample_copy)
        session.mutate(lambda p: setattr(p, "insured", "Mine"))
        sample_copy.write_text(sample_copy.read_text("utf-8") + " ", encoding="utf-8")
        session.save(force=True)
        assert load_program(sample_copy).insured == "Mine"
        session.mutate(lambda p: setattr(p, "insured", "Mine Again"))
        session.save()  # the guard was re-armed by the forced save; no raise
        assert load_program(sample_copy).insured == "Mine Again"

    def test_reload_takes_the_file_and_drops_local_edits(self, sample_copy) -> None:
        session = EditSession.open(sample_copy)
        session.mutate(lambda p: setattr(p, "insured", "Mine"))
        program = load_program(sample_copy)
        program.insured = "Theirs"
        sample_copy.write_text(dumps_program(program), encoding="utf-8")
        session.reload()
        assert session.program.insured == "Theirs"
        assert not session.dirty
        assert not session.undo()  # history belongs to the discarded edits
        session.save()  # no raise: the guard re-armed on reload


class TestInputHelpers:
    def test_share_parsing(self) -> None:
        assert parse_share_pct("35") == 3_500
        assert parse_share_pct("33.33") == 3_333
        assert parse_share_pct("100%") == BPS_SCALE
        assert parse_share_pct("0") == 0
        assert parse_share_pct("101") is None
        assert parse_share_pct("33.333") is None  # sub-bps precision rejected
        assert parse_share_pct("abc") is None

    @pytest.mark.asyncio
    async def test_carrier_suggester_prefix_and_fuzzy(self) -> None:
        suggester = CarrierSuggester(["Swiss Re", "Sompo", "Munich Re"])
        assert await suggester.get_suggestion("Swi") == "Swiss Re"
        assert await suggester.get_suggestion("swiss re corporate") == "Swiss Re"
        assert await suggester.get_suggestion("zzz") is None


class TestBrowser:
    def test_bump_stem(self) -> None:
        assert _bump_stem("atomic-2026") == "atomic-2027"
        assert _bump_stem("prog") == "prog-renewal"

    @pytest.mark.asyncio
    async def test_browser_lists_programs_with_badges(self, tmp_path, monkeypatch) -> None:
        programs = tmp_path / "programs"
        programs.mkdir()
        shutil.copy(SAMPLE, programs / "atomic-2026.json")
        monkeypatch.chdir(tmp_path)
        app = TowerkitApp()
        async with app.run_test(size=(120, 40)):
            browser = app.screen
            assert isinstance(browser, ProgramBrowser)
            table = browser.query_one("#programs")
            assert table.row_count == 1
            row = table.get_row_at(0)
            assert row[1] == "atomic-2026.json"
            assert row[2] == "Atomic Industries, Inc."
            assert "⚠" in row[0]  # the deliberate 80%-placed warning

    @pytest.mark.asyncio
    async def test_open_editor_from_browser(self, tmp_path, monkeypatch) -> None:
        programs = tmp_path / "programs"
        programs.mkdir()
        shutil.copy(SAMPLE, programs / "atomic-2026.json")
        monkeypatch.chdir(tmp_path)
        app = TowerkitApp()
        async with app.run_test(size=(120, 40)) as pilot:
            await pilot.press("enter")
            await pilot.pause()
            assert isinstance(app.screen, EditorScreen)

    @pytest.mark.asyncio
    async def test_delete_removes_the_backup_sidecar_too(self, tmp_path, monkeypatch) -> None:
        from towerkit.atomicio import backup_path

        programs = tmp_path / "programs"
        programs.mkdir()
        target = programs / "atomic-2026.json"
        shutil.copy(SAMPLE, target)
        # a prior save leaves a `.bak` sidecar beside the program
        from towerkit.model import dump_program

        program = load_program(target)
        dump_program(program, target)
        assert backup_path(target).exists()
        monkeypatch.chdir(tmp_path)
        app = TowerkitApp()
        async with app.run_test(size=(120, 40)) as pilot:
            await pilot.press("d")
            await pilot.pause()
            from towerkit.tui.widgets.modals import ConfirmModal

            assert isinstance(app.screen, ConfirmModal)
            app.screen.query_one("#yes", Button).press()
            await pilot.pause()
        assert not target.exists()
        assert not backup_path(target).exists()


class TestEditor:
    @pytest.mark.asyncio
    async def test_editor_shows_structure_and_diagnostics(self, sample_copy, monkeypatch) -> None:
        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)):
            editor = app.screen
            assert isinstance(editor, EditorScreen)
            tree = editor.query_one("#structure")
            labels = [str(n.label) for n in editor._walk(tree.root)]
            assert any("Layers (14)" in label for label in labels)
            assert any("Umbrella" in label for label in labels)
            # the 80%-placed 3rd Excess is highlighted gold, not just marked
            from rich.text import Text as RichText

            warn_nodes = [
                n.label for n in editor._walk(tree.root)
                if isinstance(n.label, RichText) and "3rd Excess" in str(n.label)
            ]
            assert warn_nodes and "#CB7E03" in str(warn_nodes[0].style)
            diags = editor.query_one("#diagnostics")
            assert len(diags.children) == 1  # the 80%-placed warning

    @pytest.mark.asyncio
    async def test_over_signing_blocked_at_input(self, sample_copy, monkeypatch) -> None:
        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            editor.selected = ("layer", "umbrella")
            await editor._rebuild_detail()
            await pilot.pause()
            share = editor.query_one("#p-share-0")
            share.value = "90"  # AIG 60→90 with Berkley at 40 would be 130%
            editor._commit_participant(share, "p-share-0")
            await pilot.pause()
            layer = editor._layer("umbrella")
            assert layer.signed_bps == BPS_SCALE  # unchanged: block, don't flag
            assert share.value == "60"  # display restored

    @pytest.mark.asyncio
    async def test_money_field_commit_updates_program(self, sample_copy, monkeypatch) -> None:
        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            editor.selected = ("layer", "umbrella")
            await editor._rebuild_detail()
            await pilot.pause()
            attach = editor.query_one("#f-layer-attach")
            attach.value = "2.5m"
            editor._commit_input(attach)
            await pilot.pause()
            assert editor._layer("umbrella").attach == 2_500_000
            # the gap this creates shows up immediately in diagnostics
            assert any(
                d.code == "line-gap" for d in editor.session.diagnostics().errors
            )

    @pytest.mark.asyncio
    async def test_undo_key(self, sample_copy, monkeypatch) -> None:
        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            editor.session.mutate(lambda p: setattr(p, "insured", "Other Co"))
            await pilot.press("u")
            await pilot.pause()
            assert editor.session.program.insured == "Atomic Industries, Inc."

    @pytest.mark.asyncio
    async def test_save_with_errors_prompts_not_blocks(self, sample_copy, monkeypatch) -> None:
        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            # seed an error: umbrella attach moves, creating a gap
            editor.session.mutate(lambda p: setattr(editor._layer("umbrella"), "attach", 3_000_000))
            before = sample_copy.read_bytes()
            await pilot.press("ctrl+s")
            await pilot.pause()
            from towerkit.tui.widgets.modals import ConfirmModal

            assert isinstance(app.screen, ConfirmModal)  # asked, not silent
            await pilot.press("escape")  # cancel
            await pilot.pause()
            assert sample_copy.read_bytes() == before  # cancelled → untouched

    @pytest.mark.asyncio
    async def test_create_valid_program_from_scratch(self, tmp_path, monkeypatch) -> None:
        """The §9 gate for the TUI, in miniature: start blank, build a valid
        program using only editor operations, save, and validate the file."""
        monkeypatch.chdir(tmp_path)
        app = TowerkitApp(new=True)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            # name the insured via the program form
            editor.selected = ("program", None)
            await editor._rebuild_detail()
            await pilot.pause()
            insured = editor.query_one("#f-insured")
            insured.value = "Scratch Built Co"
            editor._commit_input(insured)
            # add a primary layer on the default gl line (attach suggested: 0)
            editor.selected = ("layers-group", None)
            await editor.action_add_node()
            await pilot.pause()
            kind, layer_id = editor.selected
            assert kind == "layer"
            layer = editor._layer(layer_id)
            assert layer.attach == 0
            # set its limit and fill it with one carrier at 100%
            limit = editor.query_one("#f-layer-limit")
            limit.value = "5m"
            editor._commit_input(limit)
            editor.query_one("#add-participant").press()
            await pilot.pause()
            carrier = editor.query_one("#p-carrier-0")
            carrier.value = "Zurich"
            editor._commit_input(carrier)
            # add a retention on gl
            editor.selected = ("retentions-group", None)
            await editor.action_add_node()
            await pilot.pause()
            # save via the name prompt
            editor.action_save()
            await pilot.pause()
            from towerkit.tui.widgets.modals import PromptModal

            assert isinstance(app.screen, PromptModal)
            prompt = app.screen.query_one("#prompt")
            prompt.value = "scratch"
            await pilot.press("enter")
            await pilot.pause()
        saved = tmp_path / "programs" / "scratch.json"
        assert saved.exists()
        from towerkit.validate import validate_file

        program, diags = validate_file(saved)
        assert program is not None
        assert diags.ok, [str(d) for d in diags.errors]
        assert program.insured == "Scratch Built Co"
        assert program.layers[0].signed_bps == BPS_SCALE

    @pytest.mark.asyncio
    async def test_save_as_refuses_an_existing_file_name(
        self, tmp_path, monkeypatch
    ) -> None:
        """The prompt-branch of _do_save (session.path is None — Save As on
        a brand-new program) must not silently clobber a whole existing
        program file. It should refuse and leave both files untouched."""
        monkeypatch.chdir(tmp_path)
        programs_dir = tmp_path / "programs"
        programs_dir.mkdir()
        existing = programs_dir / "existing.json"
        shutil.copy(SAMPLE, existing)
        before = existing.read_bytes()

        app = TowerkitApp(new=True)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            editor.selected = ("program", None)
            await editor._rebuild_detail()
            await pilot.pause()
            insured = editor.query_one("#f-insured")
            insured.value = "Should Not Land"
            editor._commit_input(insured)
            editor.action_save()
            await pilot.pause()
            from towerkit.tui.widgets.modals import ConfirmModal, PromptModal

            if isinstance(app.screen, ConfirmModal):  # a blank program has errors
                app.screen.query_one("#yes", Button).press()
                await pilot.pause()

            assert isinstance(app.screen, PromptModal)
            prompt = app.screen.query_one("#prompt")
            prompt.value = "existing"
            await pilot.press("enter")
            await pilot.pause()
            # refused, not saved: back in the editor, nothing on disk moved
            assert isinstance(app.screen, EditorScreen)
            assert app.screen.session.path is None
        assert existing.read_bytes() == before

    @pytest.mark.asyncio
    async def test_save_over_a_changed_file_offers_a_choice(
        self, sample_copy, monkeypatch
    ) -> None:
        from towerkit.tui.widgets.modals import StaleFileModal

        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            assert isinstance(editor, EditorScreen)
            editor.session.mutate(lambda p: setattr(p, "insured", "Mine"))
            sample_copy.write_text(sample_copy.read_text("utf-8") + " ", encoding="utf-8")
            await pilot.press("ctrl+s")
            await pilot.pause()
            assert isinstance(app.screen, StaleFileModal)
            # default focus must land on the non-destructive choice: Reload
            # and Overwrite both discard someone's work, Keep editing loses
            # nothing — an unread Enter must not be able to throw work away
            assert app.screen.query_one("#keep", Button).has_focus
            await pilot.press("escape")  # keep editing
            await pilot.pause()
            assert editor.session.program.insured == "Mine"  # nothing lost


class TestAppliesToLayout:
    @pytest.mark.asyncio
    async def test_all_six_line_checkboxes_visible(self, sample_copy, monkeypatch) -> None:
        # Regression: a single horizontal row silently clipped everything
        # after the third checkbox in the 46-cell detail pane.
        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            editor.selected = ("layer", "umbrella")
            await editor._rebuild_detail()
            await pilot.pause()
            detail = editor.query_one("#detail")
            for line_id in ("gl", "al", "el", "pl", "cy", "pr"):
                box = editor.query_one(f"#applies-{line_id}")
                assert box.region.width > 0, f"{line_id} checkbox not laid out"
                assert box.region.x + box.region.width <= (
                    detail.region.x + detail.region.width
                ), f"{line_id} checkbox clipped off the detail pane"


class TestRenderOptionsMenu:
    @pytest.mark.asyncio
    async def test_menu_sets_theme_and_toggles(self, sample_copy, monkeypatch) -> None:
        # copy themes next to programs so the menu finds them
        import shutil as sh

        themes = sample_copy.parent.parent / "themes"
        themes.mkdir(exist_ok=True)
        sh.copy(REPO / "themes" / "marsh.json", themes / "marsh.json")
        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            assert editor.tower_theme.name == "default"
            assert app.show_totals and app.show_premiums
            await pilot.press("t")
            await pilot.pause()
            from towerkit.tui.widgets.modals import RenderOptionsModal

            assert isinstance(app.screen, RenderOptionsModal)
            modal = app.screen
            modal.query_one("#themes").highlighted = 1  # marsh
            modal.query_one("#opt-premiums").value = False
            modal.query_one("#apply").press()
            await pilot.pause()
            assert editor.tower_theme.name == "marsh"
            assert editor.theme_path == Path("themes/marsh.json")
            assert app.show_premiums is False
            assert app.show_totals is True


class TestPrivatePrograms:
    @pytest.mark.asyncio
    async def test_private_subdir_listed_in_browser(self, tmp_path, monkeypatch) -> None:
        programs = tmp_path / "programs"
        (programs / "private").mkdir(parents=True)
        shutil.copy(SAMPLE, programs / "atomic-2026.json")
        shutil.copy(SAMPLE, programs / "private" / "client.json")
        monkeypatch.chdir(tmp_path)
        app = TowerkitApp()
        async with app.run_test(size=(120, 40)):
            table = app.screen.query_one("#programs")
            names = [table.get_row_at(i)[1] for i in range(table.row_count)]
            assert "atomic-2026.json" in names
            assert "client.json" in names


class TestLayerNotesField:
    @pytest.mark.asyncio
    async def test_note_entered_in_form_lands_on_the_layer(self, sample_copy, monkeypatch) -> None:
        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            editor.selected = ("layer", "umbrella")
            await editor._rebuild_detail()
            await pilot.pause()
            notes = editor.query_one("#f-layer-notes")
            notes.value = "quota share under negotiation"
            editor._commit_input(notes)
            await pilot.pause()
            assert editor._layer("umbrella").notes == "quota share under negotiation"


class TestSoiDetailFields:
    @pytest.mark.asyncio
    async def test_detail_prose_lands_on_the_layer(self, sample_copy, monkeypatch) -> None:
        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            editor.selected = ("layer", "umbrella")
            await editor._rebuild_detail()
            await pilot.pause()
            limits = editor.query_one("#f-layer-limits-detail")
            limits.value = "Each Occurrence $5,000,000; Aggregate $5,000,000"
            editor._commit_input(limits)
            retention = editor.query_one("#f-layer-retention-detail")
            retention.value = "Retention $10,000"
            editor._commit_input(retention)
            await pilot.pause()
            layer = editor._layer("umbrella")
            assert layer.limits_detail == "Each Occurrence $5,000,000; Aggregate $5,000,000"
            assert layer.retention_detail == "Retention $10,000"


class TestPersistedRenderSettings:
    @pytest.mark.asyncio
    async def test_menu_choice_saves_into_the_program(self, sample_copy, monkeypatch) -> None:
        import shutil as sh

        themes = sample_copy.parent.parent / "themes"
        themes.mkdir(exist_ok=True)
        sh.copy(REPO / "themes" / "marsh.json", themes / "marsh.json")
        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            await pilot.press("t")
            await pilot.pause()
            modal = app.screen
            modal.query_one("#themes").highlighted = 1  # marsh
            modal.query_one("#opt-cell-premiums").value = True
            modal.query_one("#apply").press()
            await pilot.pause()
            stored = editor.session.program.render
            assert stored is not None
            assert stored.theme == "themes/marsh.json"
            assert stored.cell_premiums is True
            editor.session.save()
        # a fresh session opens with the same settings
        app2 = TowerkitApp(path=sample_copy)
        async with app2.run_test(size=(140, 45)):
            editor2 = app2.screen
            assert editor2.tower_theme.name == "marsh"
            assert app2.cell_premiums is True


class TestLineIdLocked:
    @pytest.mark.asyncio
    async def test_id_field_is_locked_and_label_not_polluted(self, tmp_path, monkeypatch) -> None:
        monkeypatch.chdir(tmp_path)
        app = TowerkitApp(new=True)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            editor.selected = ("lines-group", None)
            await editor.action_add_node()
            await pilot.pause()
            name = editor.query_one("#f-line-name")
            name.value = "Directors and Officers"
            editor._commit_input(name)
            await pilot.pause()
            kind, line_id = editor.selected
            assert line_id == "directors-and-officers"  # id auto-generated
            line = editor._line(line_id)
            assert line.abbr is None  # the label was NOT auto-created…
            assert line.label == "DAO"  # …it derives from the name's initials
            id_field = editor.query_one("#f-line-id")
            assert id_field.disabled  # locked, not editable


class TestPlaceholderNameCommit:
    @pytest.mark.asyncio
    async def test_blur_with_default_name_does_not_consume_id_generation(
        self, tmp_path, monkeypatch
    ) -> None:
        # Repro: add a line, let the name field blur with its untouched
        # default ("New Line"), THEN type the real name. The id must come
        # from the real name, never from the placeholder.
        monkeypatch.chdir(tmp_path)
        app = TowerkitApp(new=True)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            editor.selected = ("lines-group", None)
            await editor.action_add_node()
            await pilot.pause()
            name = editor.query_one("#f-line-name")
            editor._commit_input(name)  # blur/drain with the default value
            await pilot.pause()
            name.value = "Inland Marine"
            editor._commit_input(name)
            await pilot.pause()
            _, line_id = editor.selected
            assert line_id == "inland-marine"
            assert editor._line(line_id).name == "Inland Marine"

    @pytest.mark.asyncio
    async def test_layer_default_name_blur_keeps_id_generation(
        self, tmp_path, monkeypatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        app = TowerkitApp(new=True)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            editor.selected = ("lines-group", None)
            await editor.action_add_node()
            await pilot.pause()
            editor.selected = ("layers-group", None)
            await editor.action_add_node()
            await pilot.pause()
            name = editor.query_one("#f-layer-name")
            editor._commit_input(name)  # blur with default "New Layer"
            await pilot.pause()
            name.value = "Umbrella"
            editor._commit_input(name)
            await pilot.pause()
            _, layer_id = editor.selected
            assert layer_id == "umbrella"


class TestLineReorder:
    @pytest.mark.asyncio
    async def test_shift_arrows_reorder_columns(self, sample_copy, monkeypatch) -> None:
        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            assert editor.session.program.line_ids()[:3] == ["gl", "al", "el"]
            editor.selected = ("line", "al")
            await pilot.press("left_square_bracket")  # the real key path
            await pilot.pause()
            assert editor.session.program.line_ids()[:3] == ["al", "gl", "el"]
            # geometry follows the new order
            from towerkit.layout import build_layout

            tower = build_layout(editor.session.program)
            assert tower.columns[0].line_id == "al"
            # and it undoes like any other edit
            assert editor.session.undo()
            assert editor.session.program.line_ids()[:3] == ["gl", "al", "el"]

    @pytest.mark.asyncio
    async def test_bounds_are_safe(self, sample_copy, monkeypatch) -> None:
        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            editor.selected = ("line", "gl")
            await pilot.press("left_square_bracket")  # already first: no-op
            await pilot.pause()
            assert editor.session.program.line_ids()[0] == "gl"


class TestHelp:
    @pytest.mark.asyncio
    async def test_question_mark_opens_help(self, sample_copy, monkeypatch) -> None:
        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            await pilot.press("question_mark")
            await pilot.pause()
            from towerkit.tui.widgets.modals import HelpModal

            assert isinstance(app.screen, HelpModal)
            assert "[ / ]" in app.screen.help_text  # the reorder keys are documented
            await pilot.press("escape")
            await pilot.pause()

    @pytest.mark.asyncio
    async def test_line_form_carries_reorder_hint(self, sample_copy, monkeypatch) -> None:
        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            editor.selected = ("line", "gl")
            await editor._rebuild_detail()
            await pilot.pause()
            from textual.widgets import Label

            texts = [str(w.render()) for w in editor.query(Label)]
            assert any("moves this column left" in t for t in texts)


class TestBlurCommitRace:
    @pytest.mark.asyncio
    async def test_edit_survives_clicking_away(self, sample_copy, monkeypatch) -> None:
        """Type into a field, then move the tree selection WITHOUT pressing
        enter: the blur-commit must land on the node the form was built for,
        not the newly selected one — 'GL OpCo' must not lose its edit."""
        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            editor.selected = ("line", "gl")
            await editor._rebuild_detail()
            await pilot.pause()
            abbr = editor.query_one("#f-line-abbr")
            abbr.focus()
            await pilot.pause()
            abbr.value = "GL OpCo"
            # selection moves before the blur-commit is processed
            editor.selected = ("layer", "umbrella")
            editor._commit_input(abbr)
            await pilot.pause()
            line = editor._line("gl")
            assert line.abbr == "GL OpCo"  # spaces intact, edit not dropped


class TestColumnLabelPrefill:
    @pytest.mark.asyncio
    async def test_label_prefills_with_name(self, sample_copy, monkeypatch) -> None:
        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            # gl has an explicit abbr in the sample: shows the abbr
            editor.selected = ("line", "gl")
            await editor._rebuild_detail()
            await pilot.pause()
            assert editor.query_one("#f-line-abbr").value == "GL"
            # a line without abbr prefills with its name
            editor.session.mutate(lambda p: setattr(p.lines[0], "abbr", None))
            await editor._rebuild_detail()
            await pilot.pause()
            assert editor.query_one("#f-line-abbr").value == "General Liability"


class TestPrimaryQuickSelect:
    @pytest.mark.asyncio
    async def test_checkbox_snaps_attach_to_zero_and_back(self, sample_copy, monkeypatch) -> None:
        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            editor.selected = ("layer", "umbrella")
            await editor._rebuild_detail()
            await pilot.pause()
            box = editor.query_one("#f-layer-primary")
            assert box.value is False  # umbrella attaches at $2M
            box.value = True
            await pilot.pause()
            layer = editor._layer("umbrella")
            assert layer.attach == 0
            assert editor.query_one("#f-layer-attach").amount == 0
            box.value = False
            await pilot.pause()
            # suggested back onto the top of the gl/al/el stack beneath it
            assert layer.attach > 0


class TestStackWorkflow:
    def test_added_excess_layers_get_ordinal_names(self) -> None:
        session = EditSession.open(SAMPLE)
        layer = session.add_layer(["pl"])  # pl stack tops at $25M
        assert layer.name == "2nd Excess"  # Excess PL already sits above zero
        assert layer.attach == 25_000_000

    def test_restack_heals_gap_and_overlap(self) -> None:
        session = EditSession.open(SAMPLE)

        def seed(p):
            by_id = {ly.id: ly for ly in p.layers}
            by_id["xs-1"].attach = 30_000_000  # gap below
            by_id["xs-2"].attach = 50_000_000  # overlap above

        session.mutate(seed)
        assert not session.diagnostics().ok
        session.restack()
        diags = session.diagnostics()
        assert diags.ok, [str(d) for d in diags.errors]
        by_id = {ly.id: ly for ly in session.program.layers}
        assert by_id["xs-1"].attach == 27_000_000
        assert by_id["xs-2"].attach == 52_000_000
        assert session.undo()  # restack is one undoable edit

    def test_follows_layer_auto_heals_when_underlying_changes(self) -> None:

        session = EditSession.open(SAMPLE)

        def make_follows(p):
            by_id = {ly.id: ly for ly in p.layers}
            by_id["umbrella"].follows_underlying = True

        session.mutate(make_follows)
        umbrella = next(ly for ly in session.program.layers if ly.id == "umbrella")
        assert umbrella.attach == 2_000_000
        # grow the GL primary: the follows umbrella re-seats itself
        session.mutate(
            lambda p: setattr(
                next(ly for ly in p.layers if ly.id == "primary-gl"), "limit", 3_000_000
            )
        )
        assert umbrella.attach == 3_000_000


class TestNodeDiagnosticsDrillThrough:
    @pytest.mark.asyncio
    async def test_selected_line_shows_its_own_warning(self, tmp_path, monkeypatch) -> None:
        import json

        # a program with intentionally no retentions: every line warns
        source = json.loads(SAMPLE.read_text())
        source["retentions"] = []
        programs = tmp_path / "programs"
        programs.mkdir()
        target = programs / "no-ret.json"
        target.write_text(json.dumps(source))
        monkeypatch.chdir(tmp_path)
        app = TowerkitApp(path=target)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            editor.selected = ("line", "gl")
            await editor._rebuild_detail()
            await pilot.pause()
            from textual.widgets import Label

            texts = [str(w.render()) for w in editor.query(Label)]
            assert any("no retention recorded" in t for t in texts), texts


class TestDirtyExitOffersSave:
    async def test_escape_then_save_writes_and_exits(self, sample_copy) -> None:
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            assert isinstance(editor, EditorScreen)
            editor.session.mutate(lambda p: setattr(p, "insured", "Exit Save Co"))
            assert editor.session.dirty
            await pilot.press("escape")
            await pilot.pause()
            await pilot.press("s")  # Save & exit
            await pilot.pause()
        assert '"insured": "Exit Save Co"' in sample_copy.read_text()

    async def test_escape_then_discard_leaves_file_untouched(self, sample_copy) -> None:
        before = sample_copy.read_bytes()
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            editor.session.mutate(lambda p: setattr(p, "insured", "Nope Co"))
            await pilot.press("escape")
            await pilot.pause()
            await pilot.press("d")  # Discard and exit
            await pilot.pause()
        assert sample_copy.read_bytes() == before

    async def test_escape_save_with_errors_stays_in_editor(self, sample_copy) -> None:
        before = sample_copy.read_bytes()
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            # force a guaranteed validation error: duplicate layer ids
            editor.session.mutate(
                lambda p: setattr(p.layers[1], "id", p.layers[0].id)
            )
            assert editor.session.diagnostics().errors
            await pilot.press("escape")
            await pilot.pause()
            await pilot.press("s")
            await pilot.pause()
            assert isinstance(app.screen, EditorScreen)  # refused the exit
        assert sample_copy.read_bytes() == before


class TestSoiExport:
    @pytest.mark.asyncio
    async def test_x_exports_workbook_to_dist(self, sample_copy, monkeypatch) -> None:
        monkeypatch.chdir(sample_copy.parent.parent)
        before = sample_copy.read_bytes()
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            assert isinstance(editor, EditorScreen)
            await pilot.press("x")
            await pilot.pause()
            from towerkit.soi import default_filename

            out = Path("dist") / default_filename(editor.session.program)
            assert any(
                str(out) in n.message for n in app._notifications
            ), "expected a notification with the written path"
        assert out.exists()
        # program file untouched by export
        assert sample_copy.read_bytes() == before

    @pytest.mark.asyncio
    async def test_x_blocked_by_validation_errors(self, sample_copy, monkeypatch) -> None:
        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            # force validation errors: duplicate layer ids (same trick as dirty-exit tests)
            editor.session.mutate(
                lambda p: setattr(p.layers[1], "id", p.layers[0].id)
            )
            assert editor.session.diagnostics().errors
            await pilot.press("x")
            await pilot.pause()
            from towerkit.soi import default_filename

            out = Path("dist") / default_filename(editor.session.program)
            assert any(
                "validation error" in n.message for n in app._notifications
            ), "expected a notification mentioning validation errors"
        assert not out.exists()

    @pytest.mark.asyncio
    async def test_premiums_toggle_drops_premium_column(
        self, sample_copy, monkeypatch
    ) -> None:
        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            app.show_premiums = False  # the t-menu toggle's backing attr
            await pilot.press("x")
            await pilot.pause()
            from towerkit.soi import default_filename

            out = Path("dist") / default_filename(editor.session.program)
        from openpyxl import load_workbook

        headers = [c.value for c in load_workbook(out).active[1]]
        assert "Premium" not in headers
        assert headers[0] == "Insured"

    @pytest.mark.asyncio
    async def test_x_notifies_and_survives_write_failure(
        self, sample_copy, monkeypatch
    ) -> None:
        monkeypatch.chdir(sample_copy.parent.parent)

        def raiser(*args, **kwargs):
            raise PermissionError("locked")

        monkeypatch.setattr("towerkit.render.soi_xlsx.write_soi_workbook", raiser)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            assert isinstance(editor, EditorScreen)
            await pilot.press("x")
            await pilot.pause()
            assert app.screen is editor
            assert any(
                "export failed" in n.message for n in app._notifications
            ), "expected a notification with the export failure"
            from towerkit.soi import default_filename

            out = Path("dist") / default_filename(editor.session.program)
        assert not out.exists()


class TestRemoveLineCleansSublimits:
    @pytest.mark.asyncio
    async def test_sublimits_narrow_or_drop_with_their_line(
        self, sample_copy, monkeypatch
    ) -> None:
        # deleting a line must treat sublimits like retentions: entries
        # that only applied to it are removed, shared ones are narrowed —
        # never left with a dangling line ref (which fails validation and
        # is what made the line-transfer crash reachable)
        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            from towerkit.model import Sublimit

            editor.session.mutate(
                lambda p: p.sublimits.extend(
                    [
                        Sublimit(name="Shared", amount=100_000, applies_to=["gl", "al"]),
                        Sublimit(name="AlOnly", amount=50_000, applies_to=["al"]),
                    ]
                )
            )
            editor.selected = ("line", "al")
            await editor.action_remove_node()
            await pilot.pause()
            subs = editor.session.program.sublimits
            names = [s.name for s in subs]
            assert "AlOnly" not in names          # exclusive sublimit removed
            shared = next(s for s in subs if s.name == "Shared")
            assert shared.applies_to == ["gl"]    # shared sublimit narrowed
            assert all("al" not in s.applies_to for s in subs)


class TestAutocomplete:
    @pytest.mark.asyncio
    async def test_dropdowns_mount_for_known_records(self, sample_copy, monkeypatch) -> None:
        from textual_autocomplete import AutoComplete

        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            # program form: insured completes from programs on disk
            editor.selected = ("program", None)
            await editor._rebuild_detail()
            await pilot.pause()
            assert editor.query(AutoComplete)
            # layer form: one dropdown per participant carrier input
            editor.selected = ("layer", "umbrella")
            await editor._rebuild_detail()
            await pilot.pause()
            carriers = [
                w for w in editor.query(AutoComplete)
            ]
            assert carriers  # carrier inputs carry dropdown completion too


class TestSendLine:
    def _two_programs(self, tmp_path, monkeypatch):
        monkeypatch.chdir(tmp_path)
        target = tmp_path / "programs"
        target.mkdir()
        shutil.copy(SAMPLE, target / "src.json")
        shutil.copy(SAMPLE, target / "dst.json")
        return target / "src.json", target / "dst.json"

    @pytest.mark.asyncio
    async def test_copy_line_writes_target_additively(
        self, tmp_path, monkeypatch
    ) -> None:
        src, dst = self._two_programs(tmp_path, monkeypatch)
        from towerkit.model import load_program

        before = load_program(dst)
        app = TowerkitApp(path=src)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            editor.selected = ("line", "gl")
            await pilot.press("greater_than_sign")
            await pilot.pause()
            from towerkit.tui.widgets.modals import SendLineModal

            modal = app.screen
            assert isinstance(modal, SendLineModal)
            # pick dst.json in the option list, leave "move" unchecked
            options = modal.query_one("#send-targets")
            idx = next(
                i for i, p in enumerate(modal.targets) if p.name == "dst.json"
            )
            options.highlighted = idx
            modal.query_one("#send-confirm").press()
            await pilot.pause()
            # summary confirm modal — drive it by button press, same
            # mechanism as TestDirtyExitOffersSave, not a raw keypress
            # (ConfirmModal has no "y" binding, only Button ids yes/no)
            app.screen.query_one("#yes").press()
            await pilot.pause()
            after = load_program(dst)
            # additive: every original line/layer still present, gl grafted
            assert {ln.id for ln in before.lines} <= {ln.id for ln in after.lines}
            assert len(after.lines) == len(before.lines) + 1
            # pre-existing lines/layers serialise identically — not just
            # set-containment, the untouched target content is byte-for-byte
            after_lines_by_id = {ln.id: ln for ln in after.lines}
            for ln in before.lines:
                assert (
                    after_lines_by_id[ln.id].model_dump()
                    == ln.model_dump()
                )
            after_layers_by_id = {ly.id: ly for ly in after.layers}
            before_layer_ids = {ly.id for ly in before.layers}
            assert before_layer_ids <= set(after_layers_by_id)
            for ly in before.layers:
                assert (
                    after_layers_by_id[ly.id].model_dump()
                    == ly.model_dump()
                )
            # copy mode: source session untouched
            assert not editor.session.dirty

    @pytest.mark.asyncio
    async def test_move_is_undoable_and_source_file_untouched(
        self, tmp_path, monkeypatch
    ) -> None:
        src, dst = self._two_programs(tmp_path, monkeypatch)
        src_bytes = src.read_bytes()
        app = TowerkitApp(path=src)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            n_lines = len(editor.session.program.lines)
            editor.selected = ("line", "gl")
            await pilot.press("greater_than_sign")
            await pilot.pause()
            from towerkit.tui.widgets.modals import SendLineModal

            modal = app.screen
            assert isinstance(modal, SendLineModal)
            idx = next(
                i for i, p in enumerate(modal.targets) if p.name == "dst.json"
            )
            modal.query_one("#send-targets").highlighted = idx
            modal.query_one("#send-move").value = True
            modal.query_one("#send-confirm").press()
            await pilot.pause()
            app.screen.query_one("#yes").press()
            await pilot.pause()
            assert len(editor.session.program.lines) == n_lines - 1
            assert editor.session.undo()
            assert len(editor.session.program.lines) == n_lines
        assert src.read_bytes() == src_bytes  # source file never written

    @pytest.mark.asyncio
    async def test_malformed_target_refused_bytes_unchanged(
        self, tmp_path, monkeypatch
    ) -> None:
        src, dst = self._two_programs(tmp_path, monkeypatch)
        dst.write_text("{not json", encoding="utf-8")
        bad = dst.read_bytes()
        app = TowerkitApp(path=src)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            editor.selected = ("line", "gl")
            await pilot.press("greater_than_sign")
            await pilot.pause()
            from towerkit.tui.widgets.modals import SendLineModal

            modal = app.screen
            assert isinstance(modal, SendLineModal)
            idx = next(
                i for i, p in enumerate(modal.targets) if p.name == "dst.json"
            )
            modal.query_one("#send-targets").highlighted = idx
            modal.query_one("#send-confirm").press()
            await pilot.pause()
            await pilot.pause()  # notify() is itself a posted message; give
            # it a second idle-drain so this isn't racy under load
            assert any(
                "can't read" in n.message for n in app._notifications
            )
        assert dst.read_bytes() == bad

    @pytest.mark.asyncio
    async def test_invalid_target_refused_bytes_unchanged(
        self, tmp_path, monkeypatch
    ) -> None:
        # target parses as valid JSON/Program but fails semantic validation
        # (duplicate layer id) — refused before any transfer is attempted
        src, dst = self._two_programs(tmp_path, monkeypatch)
        text = dst.read_text(encoding="utf-8").replace(
            '"id": "primary-al"', '"id": "primary-gl"'
        )
        dst.write_text(text, encoding="utf-8")
        bad = dst.read_bytes()
        app = TowerkitApp(path=src)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            editor.selected = ("line", "gl")
            await pilot.press("greater_than_sign")
            await pilot.pause()
            from towerkit.tui.widgets.modals import SendLineModal

            modal = app.screen
            assert isinstance(modal, SendLineModal)
            idx = next(
                i for i, p in enumerate(modal.targets) if p.name == "dst.json"
            )
            modal.query_one("#send-targets").highlighted = idx
            modal.query_one("#send-confirm").press()
            await pilot.pause()
            await pilot.pause()
            assert any(
                "validation errors" in n.message for n in app._notifications
            )
            assert isinstance(app.screen, EditorScreen)  # no confirm modal
        assert dst.read_bytes() == bad

    @pytest.mark.asyncio
    async def test_requires_line_selection(self, sample_copy, monkeypatch) -> None:
        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            editor.selected = ("program", None)
            await pilot.press("greater_than_sign")
            await pilot.pause()
            assert isinstance(app.screen, EditorScreen)  # no modal
            assert any("select a line" in n.message for n in app._notifications)


class TestBrowserImport:
    async def _fill(self, pilot, value: str) -> None:
        pilot.app.screen.query_one(Input).value = value
        await pilot.press("enter")
        await pilot.pause()

    async def _submit_import_modal(
        self, pilot, *, path="", insured="", program="", inception="", expiry=""
    ) -> None:
        from towerkit.tui.widgets.modals import ImportFileModal

        modal = pilot.app.screen
        assert isinstance(modal, ImportFileModal)
        modal.query_one("#import-path").value = path
        modal.query_one("#import-insured").value = insured
        modal.query_one("#import-program").value = program
        modal.query_one("#import-inception").value = inception
        modal.query_one("#import-expiry").value = expiry
        modal.query_one("#import-confirm").press()
        await pilot.pause()

    @pytest.mark.asyncio
    async def test_w_writes_template_workbook(self, tmp_path, monkeypatch) -> None:
        monkeypatch.chdir(tmp_path)
        (tmp_path / "programs").mkdir()
        app = TowerkitApp()
        async with app.run_test(size=(140, 45)) as pilot:
            await pilot.press("w")
            await pilot.pause()
            await self._fill(pilot, "blank.xlsx")
        assert (tmp_path / "blank.xlsx").exists()

    @pytest.mark.asyncio
    async def test_w_overwrite_declined_leaves_file_unchanged(
        self, tmp_path, monkeypatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        (tmp_path / "programs").mkdir()
        target = tmp_path / "blank.xlsx"
        target.write_bytes(b"sentinel")
        app = TowerkitApp()
        async with app.run_test(size=(140, 45)) as pilot:
            await pilot.press("w")
            await pilot.pause()
            await self._fill(pilot, "blank.xlsx")
            from towerkit.tui.widgets.modals import ConfirmModal

            assert isinstance(app.screen, ConfirmModal)
            app.screen.query_one("#no").press()
            await pilot.pause()
        assert target.read_bytes() == b"sentinel"

    @pytest.mark.asyncio
    async def test_w_overwrite_confirmed_rewrites_file(
        self, tmp_path, monkeypatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        (tmp_path / "programs").mkdir()
        target = tmp_path / "blank.xlsx"
        target.write_bytes(b"sentinel")
        app = TowerkitApp()
        async with app.run_test(size=(140, 45)) as pilot:
            await pilot.press("w")
            await pilot.pause()
            await self._fill(pilot, "blank.xlsx")
            from towerkit.tui.widgets.modals import ConfirmModal

            assert isinstance(app.screen, ConfirmModal)
            app.screen.query_one("#yes").press()
            await pilot.pause()
        assert target.read_bytes() != b"sentinel"

    @pytest.mark.asyncio
    async def test_i_imports_filled_template_and_opens_editor(
        self, tmp_path, monkeypatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        (tmp_path / "programs").mkdir()
        # write_template's worked example rows import as-is (matching
        # tests/test_cli.py::TestTemplate::test_template_then_import) —
        # insured/program are supplied by the user, mirroring the CLI's
        # required --insured/--program flags (import_schedule defaults
        # both to "" and Program.insured/program require min_length=1,
        # so the browser prompts for them rather than guessing).
        src = _filled_template(tmp_path)
        app = TowerkitApp()
        async with app.run_test(size=(140, 45)) as pilot:
            await pilot.press("i")
            await pilot.pause()
            await self._submit_import_modal(
                pilot, path=str(src), insured="Example Co", program="Property"
            )
            assert isinstance(app.screen, EditorScreen)  # opened for editing
            new_files = list((tmp_path / "programs").glob("*.json"))
            assert len(new_files) == 1

    @pytest.mark.asyncio
    async def test_i_refuses_existing_output(self, tmp_path, monkeypatch) -> None:
        monkeypatch.chdir(tmp_path)
        (tmp_path / "programs").mkdir()
        src = _filled_template(tmp_path)
        app = TowerkitApp()
        async with app.run_test(size=(140, 45)) as pilot:
            await pilot.press("i")
            await pilot.pause()
            await self._submit_import_modal(
                pilot, path=str(src), insured="Example Co", program="Property"
            )
        target = next((tmp_path / "programs").glob("*.json"))
        before = target.read_bytes()
        app2 = TowerkitApp()
        async with app2.run_test(size=(140, 45)) as pilot:
            await pilot.press("i")
            await pilot.pause()
            await self._submit_import_modal(
                pilot, path=str(src), insured="Example Co", program="Property"
            )
            assert any(
                "not overwriting" in n.message for n in app2._notifications
            )
        assert target.read_bytes() == before

    @pytest.mark.asyncio
    async def test_i_bad_source_notifies(self, tmp_path, monkeypatch) -> None:
        monkeypatch.chdir(tmp_path)
        (tmp_path / "programs").mkdir()
        app = TowerkitApp()
        async with app.run_test(size=(140, 45)) as pilot:
            await pilot.press("i")
            await pilot.pause()
            await self._submit_import_modal(
                pilot,
                path=str(tmp_path / "nope.txt"),
                insured="Example Co",
                program="Property",
            )
            assert any(
                "import failed" in n.message for n in app._notifications
            )
        assert not list((tmp_path / "programs").glob("*.json"))

    @pytest.mark.asyncio
    async def test_i_row_level_error_notifies_and_writes_nothing(
        self, tmp_path, monkeypatch
    ) -> None:
        # A bad "share" cell (unlike a bad "limit") drops only that row's
        # participant, not the layer: program_from_rows creates the layer
        # first (ingest.py:264-273) and only reaches the share parse after
        # (ingest.py:283-289), so the layer survives with zero participants.
        # to_program() therefore SUCCEEDS (an unplaced layer is a warning,
        # not a validate_program error) while draft.diagnostics.errors is
        # still non-empty from the row-level rows.share error recorded
        # during parsing — the one case _finish_import's guard must catch
        # via draft.diagnostics.errors, since `program is None` alone
        # would miss it (verified empirically; a bad "limit" cell instead
        # cascades into a line-base validate_program failure and raises
        # ProgramInvalidError, which is the already-covered branch).
        monkeypatch.chdir(tmp_path)
        (tmp_path / "programs").mkdir()
        from openpyxl import load_workbook

        from towerkit.ingest_template import write_template

        src = write_template(tmp_path / "sched.xlsx")
        wb = load_workbook(src)
        ws = wb.worksheets[0]
        headers = [str(c.value or "").strip().lower() for c in ws[1]]
        ws.cell(row=2, column=headers.index("share") + 1, value="banana%")
        wb.save(src)

        app = TowerkitApp()
        async with app.run_test(size=(140, 45)) as pilot:
            await pilot.press("i")
            await pilot.pause()
            await self._submit_import_modal(
                pilot, path=str(src), insured="Example Co", program="Property"
            )
            assert isinstance(app.screen, ProgramBrowser)  # still alive, no editor
            assert any(
                "nothing written" in n.message for n in app._notifications
            )
        assert not list((tmp_path / "programs").glob("*.json")), (
            "a schedule with row-level errors must not be written"
        )

    @pytest.mark.asyncio
    async def test_i_imports_text_schedule_with_dates(
        self, tmp_path, monkeypatch
    ) -> None:
        # the dead-end case that motivated ImportFileModal: parse_tower has
        # no date syntax, so a .txt schedule can only pick up a policy
        # period when the modal's optional inception/expiry are filled.
        monkeypatch.chdir(tmp_path)
        (tmp_path / "programs").mkdir()
        src = tmp_path / "schedule.txt"
        src.write_text(self.PASTE, encoding="utf-8")
        app = TowerkitApp()
        async with app.run_test(size=(140, 45)) as pilot:
            await pilot.press("i")
            await pilot.pause()
            await self._submit_import_modal(
                pilot,
                path=str(src),
                insured="Atomic Industries",
                program="Property",
                inception="Jan 1 2026",
                expiry="1/1/2027",
            )
            assert isinstance(app.screen, EditorScreen)
            files = list((tmp_path / "programs").glob("*.json"))
            assert len(files) == 1
            from towerkit.model import load_program

            program = load_program(files[0])
            assert program.insured == "Atomic Industries"
            assert len(program.layers) == 2

    PASTE = (
        "Primary 10M — Chubb 100% — 250,000\n"
        "15M xs 10M — AXA XL 60%, Sompo 40% — 180k\n"
        "SIR 500k\n"
    )

    @pytest.mark.asyncio
    async def test_p_pastes_schedule_with_fields(self, tmp_path, monkeypatch) -> None:
        monkeypatch.chdir(tmp_path)
        (tmp_path / "programs").mkdir()
        app = TowerkitApp()
        async with app.run_test(size=(140, 45)) as pilot:
            await pilot.press("p")
            await pilot.pause()
            from towerkit.tui.widgets.modals import PasteImportModal

            modal = app.screen
            assert isinstance(modal, PasteImportModal)
            modal.query_one("#paste-text").text = self.PASTE
            modal.query_one("#paste-insured").value = "Atomic Industries"
            modal.query_one("#paste-program").value = "Property"
            modal.query_one("#paste-inception").value = "Jan 1 2026"
            modal.query_one("#paste-expiry").value = "1/1/2027"
            modal.query_one("#paste-confirm").press()
            await pilot.pause()
            assert isinstance(app.screen, EditorScreen)
            files = list((tmp_path / "programs").glob("*.json"))
            assert len(files) == 1
            from towerkit.model import load_program

            program = load_program(files[0])
            assert program.insured == "Atomic Industries"
            assert len(program.layers) == 2

    @pytest.mark.asyncio
    async def test_p_empty_text_dismisses_as_noop(self, tmp_path, monkeypatch) -> None:
        monkeypatch.chdir(tmp_path)
        (tmp_path / "programs").mkdir()
        app = TowerkitApp()
        async with app.run_test(size=(140, 45)) as pilot:
            await pilot.press("p")
            await pilot.pause()
            from towerkit.tui.widgets.modals import PasteImportModal

            assert isinstance(app.screen, PasteImportModal)
            app.screen.query_one("#paste-confirm").press()
            await pilot.pause()
            assert isinstance(app.screen, ProgramBrowser)
            assert not any(n.severity == "error" for n in app._notifications)
        assert not list((tmp_path / "programs").glob("*.json"))

    @pytest.mark.asyncio
    async def test_p_unparseable_schedule_notifies_and_writes_nothing(
        self, tmp_path, monkeypatch
    ) -> None:
        # this exercises _finish_import's ProgramInvalidError path (a
        # diagnostics-only draft); parse_tower never raises, so
        # action_paste_import's except-Exception wrapper is defensive
        # only, untested here.
        monkeypatch.chdir(tmp_path)
        (tmp_path / "programs").mkdir()
        app = TowerkitApp()
        async with app.run_test(size=(140, 45)) as pilot:
            await pilot.press("p")
            await pilot.pause()
            from towerkit.tui.widgets.modals import PasteImportModal

            modal = app.screen
            assert isinstance(modal, PasteImportModal)
            modal.query_one("#paste-text").text = "not a schedule at all\n"
            modal.query_one("#paste-insured").value = "Atomic Industries"
            modal.query_one("#paste-program").value = "Property"
            modal.query_one("#paste-confirm").press()
            await pilot.pause()
            assert isinstance(app.screen, ProgramBrowser)
            assert any(
                "import failed" in n.message for n in app._notifications
            )
        assert not list((tmp_path / "programs").glob("*.json"))


class TestSchematicToggle:
    @pytest.mark.asyncio
    async def test_toggle_persists_and_export_honors_it(
        self, sample_copy, monkeypatch
    ) -> None:
        from openpyxl import load_workbook

        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            await pilot.press("t")
            await pilot.pause()
            modal = app.screen
            modal.query_one("#opt-soi-schematic").value = True
            modal.query_one("#apply").press()
            await pilot.pause()
            stored = editor.session.program.render
            assert stored is not None and stored.soi_schematic is True
            assert app.soi_schematic is True
            editor.session.save()
            await pilot.press("x")  # export honors the toggle
            await pilot.pause()
            from towerkit.soi import default_filename

            out = Path("dist") / default_filename(editor.session.program)
            assert out.exists()
            assert len(load_workbook(out).sheetnames) == 2
        # a fresh session reopens with the toggle set
        app2 = TowerkitApp(path=sample_copy)
        async with app2.run_test(size=(140, 45)):
            assert app2.soi_schematic is True


class TestRenameAlwaysFollowsTheId:
    """Grant's typo trap: the id used to regenerate ONLY while it still looked
    like a placeholder ('line-2'). The first committed name burned that one
    chance, so a typo in it was permanent and the id field is disabled — there
    was no way back. Renaming now always re-derives the id and cascades."""

    @pytest.mark.asyncio
    async def test_fixing_a_typo_updates_a_settled_line_id(
        self, tmp_path, monkeypatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        app = TowerkitApp(new=True)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            editor.selected = ("lines-group", None)
            await editor.action_add_node()
            await pilot.pause()

            name = editor.query_one("#f-line-name")
            name.value = "Genral Liability"          # the typo
            editor._commit_input(name)
            await pilot.pause()
            _, typo_id = editor.selected
            assert typo_id == "genral-liability"     # settled — no longer a placeholder

            name = editor.query_one("#f-line-name")
            name.value = "General Liability"         # the fix
            editor._commit_input(name)
            await pilot.pause()

            _, fixed_id = editor.selected
            assert fixed_id == "general-liability"
            assert editor._line(fixed_id) is not None
            assert editor._line(typo_id) is None

    @pytest.mark.asyncio
    async def test_renaming_a_line_cascades_into_everything_pointing_at_it(
        self, tmp_path, monkeypatch
    ) -> None:
        """appliesTo on layers, retentions and sublimits all hold line ids."""
        monkeypatch.chdir(tmp_path)
        app = TowerkitApp(new=True)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            program = editor.session.program
            old_id = program.lines[0].id
            assert any(old_id in ly.applies_to for ly in program.layers) or True

            editor.session.add_layer([old_id])
            await pilot.pause()

            editor.selected = ("line", old_id)
            await editor._rebuild_detail()
            await pilot.pause()

            name = editor.query_one("#f-line-name")
            name.value = "Excess Casualty"
            editor._commit_input(name)
            await pilot.pause()

            program = editor.session.program
            new_id = "excess-casualty"
            assert [ln.id for ln in program.lines if ln.id == new_id] == [new_id]
            for layer in program.layers:
                assert old_id not in layer.applies_to
            for retention in program.retentions:
                assert old_id not in retention.applies_to
            for sublimit in program.sublimits:
                assert old_id not in sublimit.applies_to
            assert any(new_id in ly.applies_to for ly in program.layers)

    @pytest.mark.asyncio
    async def test_fixing_a_typo_updates_a_settled_layer_id(
        self, tmp_path, monkeypatch
    ) -> None:
        """The layer path carried the identical guard."""
        monkeypatch.chdir(tmp_path)
        app = TowerkitApp(new=True)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            layer = editor.session.add_layer()
            await pilot.pause()

            editor.selected = ("layer", layer.id)
            await editor._rebuild_detail()
            await pilot.pause()

            name = editor.query_one("#f-layer-name")
            name.value = "Frist Excess"              # the typo
            editor._commit_input(name)
            await pilot.pause()
            settled = editor.session.program.layers[-1].id
            assert settled == "frist-excess"

            editor.selected = ("layer", settled)
            await editor._rebuild_detail()
            await pilot.pause()
            name = editor.query_one("#f-layer-name")
            name.value = "First Excess"              # the fix
            editor._commit_input(name)
            await pilot.pause()

            assert editor.session.program.layers[-1].id == "first-excess"

    @pytest.mark.asyncio
    async def test_a_rename_does_not_collide_with_its_own_id(
        self, tmp_path, monkeypatch
    ) -> None:
        """unique_id's taken-set includes the entity being renamed, so without
        excluding self a name that re-slugs to the SAME id would drift to
        'x-2' on every edit."""
        monkeypatch.chdir(tmp_path)
        app = TowerkitApp(new=True)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            editor.selected = ("lines-group", None)
            await editor.action_add_node()
            await pilot.pause()
            name = editor.query_one("#f-line-name")
            name.value = "Cyber"
            editor._commit_input(name)
            await pilot.pause()
            assert editor.selected[1] == "cyber"

            # a cosmetic edit that slugs to the same id must keep the id
            name = editor.query_one("#f-line-name")
            name.value = "CYBER"
            editor._commit_input(name)
            await pilot.pause()
            assert editor.selected[1] == "cyber"
            assert editor._line("cyber").name == "CYBER"


class TestSaveFailures:
    @pytest.mark.asyncio
    async def test_an_unwritable_file_notifies_instead_of_killing_the_app(
        self, sample_copy, monkeypatch
    ) -> None:
        monkeypatch.chdir(sample_copy.parent.parent)
        sample_copy.chmod(0o444)
        try:
            app = TowerkitApp(path=sample_copy)
            async with app.run_test(size=(140, 45)) as pilot:
                screen = app.screen
                assert isinstance(screen, EditorScreen)
                screen.session.mutate(lambda p: setattr(p, "insured", "Mine"))
                await pilot.press("ctrl+s")
                await pilot.pause()

                assert app.is_running
                assert screen.session.dirty  # the edit is still here
        finally:
            sample_copy.chmod(0o644)

    @pytest.mark.asyncio
    async def test_a_deleted_file_is_written_back_without_a_question(
        self, sample_copy, monkeypatch
    ) -> None:
        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            screen = app.screen
            assert isinstance(screen, EditorScreen)
            screen.session.mutate(lambda p: setattr(p, "insured", "Mine"))
            sample_copy.unlink()  # git checkout, a rename, a sync client

            await pilot.press("ctrl+s")
            await pilot.pause()

            assert app.is_running
            assert sample_copy.exists()
            assert load_program(sample_copy).insured == "Mine"

    @pytest.mark.asyncio
    async def test_reload_of_a_vanished_file_does_not_kill_the_app(
        self, sample_copy, monkeypatch
    ) -> None:
        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            screen = app.screen
            assert isinstance(screen, EditorScreen)
            screen.session.mutate(lambda p: setattr(p, "insured", "Mine"))

            # the StaleFileModal's own reload branch, reached directly: the
            # file is gone by the time the user answers the question
            sample_copy.unlink()
            screen._reload_guarded()
            await pilot.pause()

            assert app.is_running
            assert screen.session.program.insured == "Mine"

    @pytest.mark.asyncio
    async def test_save_as_of_a_new_program_survives_an_unwritable_destination(
        self, tmp_path, monkeypatch
    ) -> None:
        """_do_save's on_name closure (session.path is None — Save As on a
        brand-new program) has its own `except OSError`, separate from
        _save_guarded's. Drive it through the real PromptModal, the same way
        test_save_as_refuses_an_existing_file_name does, rather than calling
        the nested on_name closure directly — it isn't a reachable attribute."""
        from towerkit.tui.widgets.modals import ConfirmModal, PromptModal

        monkeypatch.chdir(tmp_path)
        programs_dir = tmp_path / "programs"
        programs_dir.mkdir()
        programs_dir.chmod(0o555)  # read + execute, no write
        try:
            app = TowerkitApp(new=True)
            async with app.run_test(size=(140, 45)) as pilot:
                editor = app.screen
                assert isinstance(editor, EditorScreen)
                editor.selected = ("program", None)
                await editor._rebuild_detail()
                await pilot.pause()
                insured = editor.query_one("#f-insured")
                insured.value = "Mine"
                editor._commit_input(insured)
                editor.action_save()
                await pilot.pause()

                if isinstance(app.screen, ConfirmModal):  # a blank program has errors
                    app.screen.query_one("#yes", Button).press()
                    await pilot.pause()

                assert isinstance(app.screen, PromptModal)
                prompt = app.screen.query_one("#prompt")
                prompt.value = "newprogram"
                await pilot.press("enter")
                await pilot.pause()

                assert app.is_running
                assert isinstance(app.screen, EditorScreen)
                assert app.screen.session.path is None  # never landed
                assert app.screen.session.program.insured == "Mine"  # edit intact
                assert any(
                    "could not save" in n.message for n in app._notifications
                )
        finally:
            programs_dir.chmod(0o755)

    @pytest.mark.asyncio
    async def test_stale_file_overwrite_survives_a_failing_write(
        self, sample_copy, monkeypatch
    ) -> None:
        """The StaleFileModal's "overwrite" choice has its own
        `except OSError` around the forced save. Reach the modal the way
        test_save_over_a_changed_file_offers_a_choice does (file changed,
        not deleted, so this is the question-modal path, not the
        write-back-without-asking path), then make the forced write itself
        fail by monkeypatching the atomic write primitive session.py calls."""
        from towerkit.tui.widgets.modals import StaleFileModal

        def fail(*args: object, **kwargs: object) -> None:
            raise OSError(errno.EACCES, "Permission denied")

        monkeypatch.setattr("towerkit.tui.session.atomic_write_text", fail)
        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            editor = app.screen
            assert isinstance(editor, EditorScreen)
            editor.session.mutate(lambda p: setattr(p, "insured", "Mine"))
            sample_copy.write_text(sample_copy.read_text("utf-8") + " ", encoding="utf-8")

            await pilot.press("ctrl+s")
            await pilot.pause()
            assert isinstance(app.screen, StaleFileModal)

            await pilot.press("o")  # overwrite
            await pilot.pause()

            assert app.is_running
            assert isinstance(app.screen, EditorScreen)
            assert editor.session.program.insured == "Mine"  # edit intact
            assert any(
                "could not save" in n.message for n in app._notifications
            )


class TestExitGuards:
    @pytest.mark.asyncio
    async def test_esc_does_not_discard_text_still_sitting_in_a_field(
        self, sample_copy, monkeypatch
    ) -> None:
        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            screen = app.screen
            assert isinstance(screen, EditorScreen)
            field = screen.query_one("#f-insured", Input)
            field.focus()
            await pilot.pause()
            field.value = "Acme Holdings"

            # dirty is still False here: the text has not reached the model
            assert not screen.session.dirty

            await pilot.press("escape")
            await pilot.pause()

            assert app.is_running, "esc left the editor with typed text unsaved"
            assert screen.session.program.insured == "Acme Holdings"

    @pytest.mark.asyncio
    async def test_ctrl_q_asks_before_discarding_unsaved_edits(
        self, sample_copy, monkeypatch
    ) -> None:
        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            screen = app.screen
            assert isinstance(screen, EditorScreen)
            screen.session.mutate(lambda p: setattr(p, "insured", "Mine"))

            await pilot.press("ctrl+q")
            await pilot.pause()

            assert app.is_running, "ctrl+q quit past the unsaved-changes prompt"
            assert "ExitChoiceModal" in [type(s).__name__ for s in app.screen_stack]

    @pytest.mark.asyncio
    async def test_ctrl_q_still_quits_when_there_is_nothing_to_lose(
        self, sample_copy, monkeypatch
    ) -> None:
        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            assert isinstance(app.screen, EditorScreen)

            await pilot.press("ctrl+q")
            await pilot.pause()

            assert not app.is_running

    @pytest.mark.asyncio
    async def test_ctrl_q_does_not_discard_text_still_sitting_in_a_field(
        self, sample_copy, monkeypatch
    ) -> None:
        monkeypatch.chdir(sample_copy.parent.parent)
        app = TowerkitApp(path=sample_copy)
        async with app.run_test(size=(140, 45)) as pilot:
            screen = app.screen
            assert isinstance(screen, EditorScreen)
            field = screen.query_one("#f-insured", Input)
            field.focus()
            await pilot.pause()
            field.value = "Acme Holdings"

            # dirty is still False here: the text has not reached the model
            assert not screen.session.dirty

            await pilot.press("ctrl+q")
            await pilot.pause()

            assert app.is_running, "ctrl+q left the editor with typed text unsaved"
            assert screen.session.program.insured == "Acme Holdings"

    @pytest.mark.asyncio
    async def test_ctrl_q_quits_from_the_browser_too(
        self, tmp_path, monkeypatch
    ) -> None:
        monkeypatch.chdir(tmp_path)
        (tmp_path / "programs").mkdir()
        app = TowerkitApp()
        async with app.run_test(size=(140, 45)) as pilot:
            assert isinstance(app.screen, ProgramBrowser)

            await pilot.press("ctrl+q")
            await pilot.pause()

            assert not app.is_running
