"""The populate-and-reimport workbook: one registry drives export and read.

Headers stay exactly the canonical field names so a filled template imports
with zero mapping. Required-ness is conveyed by styling, never by mangling
the header text.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from .ingest import CANONICAL_FIELDS


@dataclass(frozen=True)
class ColumnSpec:
    key: str
    required: bool
    example: str


_EXAMPLES: dict[str, tuple[bool, str]] = {
    "layer": (True, "Primary"),
    "line": (False, "Property"),
    "limit": (True, "10M"),
    "attachment": (True, "0"),
    "carrier": (True, "Chubb"),
    "share": (False, "100%"),
    "premium": (False, "250,000"),
    "inception": (True, "2026-10-01"),
    "expiry": (True, "2027-10-01"),
    "policy_number": (False, ""),
}

COLUMNS: tuple[ColumnSpec, ...] = tuple(
    ColumnSpec(key, *_EXAMPLES[key]) for key in CANONICAL_FIELDS
)

_EXAMPLE_ROWS: tuple[tuple[str, ...], ...] = (
    ("Primary", "Property", "10M", "0", "Chubb", "100%", "250,000",
     "2026-10-01", "2027-10-01", ""),
    ("15M xs 10M", "Property", "15M", "10M", "AXA XL", "60", "180k",
     "2026-10-01", "2027-10-01", ""),
    ("15M xs 10M", "Property", "15M", "10M", "Sompo", "40", "",
     "2026-10-01", "2027-10-01", ""),
)


def write_template(path: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Program"
    required_fill = PatternFill("solid", fgColor="F9E6A0")
    for col, spec in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=spec.key)
        cell.font = Font(bold=True)
        if spec.required:
            cell.fill = required_fill
        ws.column_dimensions[cell.column_letter].width = max(12, len(spec.key) + 2)
    for row in _EXAMPLE_ROWS:
        ws.append(list(row))
    ws.freeze_panes = "A2"
    wb.save(path)
    return path


def read_rows(path: Path) -> list[dict[str, object]]:
    """Strict reader: headers must be canonical names (case-insensitive,
    stripped). Fuzzy header mapping is the caller's job, not towerkit's."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            return []
        keys = [str(h or "").strip().lower() for h in header]
        unknown = [k for k in keys if k and k not in CANONICAL_FIELDS]
        if unknown:
            raise ValueError(
                f"unknown columns {unknown!r}; expected {list(CANONICAL_FIELDS)!r}"
            )
        out: list[dict[str, object]] = []
        for raw in rows_iter:
            row = {
                # read-only sheets yield ragged rows; short rows are fine
                k: v
                for k, v in zip(keys, raw, strict=False)
                if k and v is not None and v != ""
            }
            if row:
                out.append(row)
        return out
    finally:
        wb.close()
