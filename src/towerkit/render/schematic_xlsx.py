"""The tower as a worksheet: merged ranges over a quantized grid.

Geometry comes from layout.py — THE SAME TowerLayout the graphic renderer
consumes — so stacking, spans and proportions match the chart by
construction. This module only quantizes the layout's normalized
coordinates onto sheet rows/columns and paints merged ranges; no tower math
happens in cell space. Colors/fonts come from the Theme, label text from
render/labels.py — both shared with the graphic.

Quantization is per-BOUNDARY, not per-block: every distinct edge float in
the layout snaps to one integer row/column, and blocks look their edges up.
Neighbours share bit-identical edge floats (layout.py's exactness rule), so
merges tile with no gaps and no overlaps by construction.

HYBRID deviation from pure proportionality (Task 9, Grant's Excel review):
`quantize_boundaries` accepts `label_spans` — (y_lo, y_hi, min_rows) triples
naming every layer's and retention's row band that carries rendered text —
and floors each to at least ITS OWN `min_rows` even when gamma-compressed
proportionality alone would round it thinner (an extreme primary deep
under a tall tower can still hit this after compression). `min_rows` is
label-aware (Task 11, Grant's second Excel review): `label_row_floor`
derives it from the span's actual rendered LINE COUNT, so a 3-line primary
heading gets more room than a 1-line share label — see `label_row_floor`
and `_label_spans`. The floor only pushes the span's upper boundary and
everything strictly after it forward; it never reopens an earlier
boundary, so tiling stays gap-free and overlap-free by construction — see
`quantize_boundaries` for the mechanism.

Narrow merges get a parallel hybrid: below `NARROW_MERGE_UNITS_PER_LINE`
Excel-width-units per text line, `wrap_text` has nowhere good to break a
word and degenerates toward one-character-per-line, so those cells switch
to `shrink_to_fit` with a shorter label pulled from the labels.py compact
ladder instead (see `_fit_label`). This is deliberately ORTHOGONAL to the
row floor above: the row floor always sizes for the full (uncompacted)
candidate line count, never the width-driven compacted one.

LINE SPACERS (Task 11, adopted improvement): layout.py's `_columns` closes
the interior gutter between columns sharing a group bucket (both ungrouped
counts as one bucket, per its own docstring) so a blanket layer spanning
several lines draws as ONE continuous shape — correct for the graphic, but
it means the schematic's grid has NO visual separation between adjacent
lines of cover whenever they're joined, and only a proportionally-scaled
(so, at high line counts, imperceptibly thin) gap when they're not. Since
layout.py's x-coordinates are shared with the graphic and must NOT change
for this, `line_column_slots` injects a schematic-only spacer COLUMN
between every pair of adjacent `layout.columns`, sized to a fixed
`SPACER_UNITS` regardless of the tower's own scale — mirroring the
graphic's inter-column GUTTER, just implemented one layer up, in the sheet
grid rather than in tower coordinates. A genuinely continuous (blanket)
rect that spans the transition still merges straight through the injected
column, so it reads unbroken; two independent per-line rects that happen
to meet exactly at the transition now have a visible gutter between them.
See `line_column_slots` for the open/close dual-mapping this requires.

AXIS LINES, GRIDLINES, FOOTER (Task 11): `_axis_lines` draws the real Y
(left, column A) and X (bottom, baseline row) axis borders in the theme's
ink colour; `_gridlines` adds a faint hairline top border at every
attachment boundary, but only through cells no block already occupies (a
line through a filled block would slice it, and the boundary is already
implied where blocks abut); `_footer` writes a small dim provenance line
below the tower. All three run AFTER every block/retention merge is
styled (`_add_border_edge` merges a new edge into whatever a cell already
has, rather than overwriting it) — same ordering rule as the merge-before-
style fix elsewhere in this module."""

from __future__ import annotations

import math
from collections.abc import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .. import __version__
from ..layout import LayerBlock, ParticipantBlock, Rect, TowerLayout, build_layout
from ..model import Program
from ..money import format_money_compact
from ..scale import DEFAULT_GAMMA
from ..theme import Chrome, Theme, contrast_text

# Module-qualified on purpose: `is_pending` is already a local and a
# parameter name in this file, and a bare import of that name is shadowed
# for the whole enclosing function — which is an UnboundLocalError at the
# call, not a name clash at import. Same trap bit mpl_program.py.
from . import labels
from .labels import (
    block_premium_label,
    carrier_only_label,
    group_label,
    heading_blocks,
    layer_heading,
    participant_label,
    retention_label,
    unplaced_label,
)
from .table_xlsx import _argb, sanitize_sheet_title

TOTAL_ROWS = 100         # quantization target across the full y-span
GRID_ROW_HEIGHT = 4.0    # uniform thin rows: ~100 × 4pt ≈ one screen of tower
AXIS_COL = 1             # column A: attachment boundaries as money labels
AXIS_WIDTH = 10.0

# Total tower width in Excel column-width units; ~columns A-AM at default
# width, per Grant's review 2026-08-13 ("the tower currently occupies
# roughly columns A-R and looks cramped ... use the full working width"),
# narrowed slightly in the follow-up review (2026-08-13, Task 11: "each
# line of cover slightly less wide"). Every render fills to this SAME
# total, regardless of how many line columns it has: AXIS_WIDTH stays
# fixed, and the tower's columns (plus the Task-11 line spacers, see
# SPACER_UNITS) are scaled by whatever per-unit rate makes their sum fill
# the remainder — same relative proportions as a fixed rate would give,
# just wider.
CANVAS_WIDTH_UNITS = 240.0
FIRST_GRID_COL = 2
TITLE_ROW, GROUP_ROW, LINE_ROW = 1, 2, 3
FIRST_GRID_ROW = 4
ROW_FLOOR = 2  # absolute safety-net minimum every labeled row band gets,
# under whatever label_row_floor(line_count) computes (belt-and-braces:
# label_row_floor already exceeds this for any real label, see below)

# Below this many Excel width units per text LINE, Excel's wrap_text has no
# good place to break a word and degenerates toward one-character-per-line
# (Grant's Excel review). Derivation: 1 Excel width unit is treated as ~= 1
# character of the sheet's font, so 2.5 units is "enough room for a short
# word or a percentage" per wrapped line — below that, shrink_to_fit with a
# shorter label reads better than a wrap attempt.
NARROW_MERGE_UNITS_PER_LINE = 2.5

# LINE_ROW header text-line height estimate for the bold 9pt font used
# there — same "characters ~= width units" heuristic as table_xlsx.py's
# _row_height, not a precise font metric.
HEADER_LINE_HEIGHT = 12.0
HEADER_TWO_LINE_HEIGHT = 2 * HEADER_LINE_HEIGHT

# Participant/retention block label line-height estimate (8pt/7pt block
# fonts) — same heuristic family as HEADER_LINE_HEIGHT, one size down.
# Used by label_row_floor (Task 11's label-aware row floor, Grant's second
# Excel review: "a 3-line primary label must render with visible breathing
# room").
LABEL_LINE_HEIGHT_PT = 11.0

# Fixed-width gutter column injected between every pair of adjacent lines
# of cover (layout.columns), independent of the tower's own scale — see
# the module docstring's "LINE SPACERS" section and `line_column_slots`.
SPACER_UNITS = 1.2

# Row gap between the tower's lowest drawn row (the retention band, if any
# — otherwise the lowest layer) and the provenance footer line.
FOOTER_GAP_ROWS = 3


def label_row_floor(label_lines: int) -> int:
    """Minimum grid-row span for a labeled block: enough GRID_ROW_HEIGHT
    rows to hold `label_lines` lines of text at LABEL_LINE_HEIGHT_PT, plus
    one row of padding (~half a line) so the label doesn't sit flush
    against the block's own edges (Grant's Excel review: thin primary
    blocks with a 3-line label rendered tight). Never below the absolute
    ROW_FLOOR safety net."""
    lines = max(label_lines, 1)
    return max(math.ceil(lines * LABEL_LINE_HEIGHT_PT / GRID_ROW_HEIGHT) + 1, ROW_FLOOR)


def quantize_boundaries(
    ys: Sequence[float],
    total_rows: int = TOTAL_ROWS,
    *,
    label_spans: Sequence[tuple[float, float, int]] = (),
) -> dict[float, int]:
    """Each distinct boundary → an integer row index, proportional over the
    full span, strictly increasing (the spec's ceil/min-1 floor: gamma
    compression already keeps small layers visible, so bumps are rare).

    `label_spans` names (y_lo, y_hi, min_rows) triples — every layer's and
    retention's row band that carries rendered text — each additionally
    floored to `rows[y_hi] - rows[y_lo] >= min_rows`, on top of the strict-
    monotonic 1-row floor every boundary already gets. `min_rows` rides
    PER SPAN (label-aware, see `label_row_floor`) rather than one module
    constant: a 3-line primary heading needs more room than a 1-line share
    label. The bump only pushes y_hi (and, via the running `prev`,
    everything strictly after it) forward — it never revisits an earlier
    boundary — so adjacent spans still tile with no gaps and no overlaps
    by construction (the module docstring's "hybrid" case)."""
    distinct = sorted(set(ys))
    if len(distinct) < 2:
        return dict.fromkeys(distinct, 0)
    lo, span = distinct[0], distinct[-1] - distinct[0]
    floors_ending_at: dict[float, list[tuple[float, int]]] = {}
    for y_lo, y_hi, min_rows in label_spans:
        floors_ending_at.setdefault(y_hi, []).append((y_lo, min_rows))
    out: dict[float, int] = {}
    prev = -1
    for y in distinct:
        row = max(round((y - lo) / span * total_rows), prev + 1)
        for y_lo, min_rows in floors_ending_at.get(y, ()):
            if y_lo in out:
                row = max(row, out[y_lo] + min_rows)
        out[y] = row
        prev = row
    return out


def _participant_lines(
    block: ParticipantBlock,
    owner: LayerBlock,
    *,
    is_heading: bool,
    follows: bool,
    is_pending: bool,
    show_premiums: bool,
) -> list[str]:
    """Candidate label lines for a participant block's anchor cell, BEFORE
    `_fit_label`'s width-driven narrow-merge compaction. Used both to
    render the cell and (via `label_row_floor`, in `_label_spans`) to size
    its row band — the row floor always sizes for this full candidate,
    never the compacted one (the row floor and the narrow-merge fix stay
    orthogonal, per the module docstring)."""
    lines: list[str] = []
    if is_heading:
        lines.append(layer_heading(owner, follows=follows))
    if block.carrier is None:
        lines.append(unplaced_label(block.share_bps, pending=is_pending))
    else:
        lines.append(participant_label(block.carrier, block.share_bps))
        if show_premiums:
            premium = block_premium_label(owner.premium, block.share_bps)
            if premium is not None:
                lines.append(premium)
    return lines


def _label_spans(
    layout: TowerLayout, program: Program, *, show_premiums: bool = True
) -> list[tuple[float, float, int]]:
    """Every row band that will actually carry rendered text, each with its
    OWN label-aware row floor (see `label_row_floor`): every participant
    block's ANCHOR rect (the one add_schematic_sheet writes text into,
    deliberately NOT LayerBlock.y0/y1 — a follows-underlying layer's
    outline is only nominally that tall; its stepped-bottom runs, what's
    actually drawn, can be a fraction of it) and each retention's first
    rect (always a single line — see retention_label). The line count for
    each participant span comes from `_participant_lines`, the SAME
    candidate-line assembly the main render loop uses, so the floor always
    matches what will actually be drawn."""
    follows = {ly.id for ly in program.layers if ly.follows_underlying}
    pending = {ly.layer_id for ly in layout.layers if labels.is_pending(ly)}
    headings = heading_blocks(layout.participants)
    layer_by_id = {ly.layer_id: ly for ly in layout.layers}
    spans: list[tuple[float, float, int]] = []
    for index, block in enumerate(layout.participants):
        anchor = max(block.rects, key=lambda r: r.width, default=None)
        if anchor is None:
            continue
        owner = layer_by_id[block.layer_id]
        lines = _participant_lines(
            block, owner,
            is_heading=headings.get(block.layer_id) == index,
            follows=block.layer_id in follows,
            is_pending=block.layer_id in pending,
            show_premiums=show_premiums,
        )
        spans.append((anchor.y0, anchor.y1, label_row_floor(len(lines))))
    spans.extend(
        (ret.rects[0].y0, ret.rects[0].y1, label_row_floor(1))
        for ret in layout.retentions if ret.rects
    )
    # the caret band carries text, so quantization must not collapse it
    spans.extend(
        (band.y0, band.y1, label_row_floor(1)) for band in layout.chevrons
    )
    return spans


def x_boundaries(layout: TowerLayout) -> tuple[float, ...]:
    """Every distinct vertical edge: column drawing extents plus every
    participant/retention rect edge (share boundaries split a line into
    share-proportional sub-columns — the 'column group' per line). Group
    bands use nominal column edges, which coincide with drawing extents at
    band boundaries (gutters only close WITHIN a group, layout.py:231-254),
    so they are covered too."""
    edges: set[float] = set()
    for column in layout.columns:
        edges.add(column.ex0)
        edges.add(column.ex1)
    for block in layout.participants:
        for rect in block.rects:
            edges.add(rect.x0)
            edges.add(rect.x1)
    for retention in layout.retentions:
        for rect in retention.rects:
            edges.add(rect.x0)
            edges.add(rect.x1)
    for band in layout.chevrons:
        edges.add(band.x0)
        edges.add(band.x1)
    return tuple(sorted(edges))


def y_boundaries(layout: TowerLayout) -> tuple[float, ...]:
    """Every distinct horizontal edge, including the zero line, every
    ref-line (= every attachment breakpoint) and the retention band."""
    ys: set[float] = {0.0}
    ys.update(y for _, y in layout.ref_lines)
    for layer in layout.layers:
        for rect in layer.outlines:
            ys.add(rect.y0)
            ys.add(rect.y1)
    for block in layout.participants:
        for rect in block.rects:
            ys.add(rect.y0)
            ys.add(rect.y1)
    for retention in layout.retentions:
        for rect in retention.rects:
            ys.add(rect.y0)
            ys.add(rect.y1)
    for band in layout.chevrons:
        ys.add(band.y0)
        ys.add(band.y1)
    return tuple(sorted(ys))


def _line_transitions(layout: TowerLayout) -> tuple[set[float], set[float]]:
    """Every x value marking the boundary between two adjacent lines of
    cover (`layout.columns`), split into `joined` (flush — layout.py
    closed the gutter, same group bucket) and `gapped` (a real layout.py
    GUTTER already separates them). See the module docstring's "LINE
    SPACERS": `joined` boundaries need a spacer COLUMN injected
    (`line_column_slots`); `gapped` boundaries already have a slot, which
    just needs its WIDTH fixed instead of scaled (the caller in
    `add_schematic_sheet`) — proportional scaling could round a gutter
    imperceptibly thin at high line counts."""
    columns = layout.columns
    joined: set[float] = set()
    gapped: set[float] = set()
    for i in range(len(columns) - 1):
        if columns[i].ex1 == columns[i + 1].ex0:
            joined.add(columns[i].ex1)
        else:
            gapped.add(columns[i].ex1)
    return joined, gapped


def line_column_slots(
    layout: TowerLayout, xs: Sequence[float]
) -> tuple[dict[float, int], dict[float, int]]:
    """Map each `x_boundaries()` value to a spreadsheet column, injecting
    one spacer slot at every JOINED line-of-cover transition (see
    `_line_transitions`) — flush column pairs get no visual separation
    from layout.py today, so a schematic-only gutter is added here
    instead. Non-joined (gapped) pairs already have a real interval
    between them; that slot is left alone here and instead given a fixed
    `SPACER_UNITS` width by the caller, same visual result, no extra slot
    needed.

    Returns `(col_of, col_of_close)`. They agree everywhere except at a
    joined-transition value, where the two rects meeting there need
    DIFFERENT integers: `col_of[x]` is right for an OPENING edge (rect.x0:
    c0 = col_of[x]) — it steps past the injected spacer. `col_of_close[x]`
    is right for a CLOSING edge, used as `col_of_close[x] - 1` (rect.x1:
    c1 = col_of_close[x] - 1) — the spacer never shifts what's already the
    last real cell of the column ending at x. A genuinely continuous
    (blanket) rect spans OUTER boundaries only, never an interior joined
    value, so it merges straight through the injected slot regardless."""
    joined, _ = _line_transitions(layout)
    col_of: dict[float, int] = {}
    col_of_close: dict[float, int] = {}
    slot = FIRST_GRID_COL
    for x in xs:
        col_of_close[x] = slot
        if x in joined:
            col_of[x] = slot + 1
            slot += 2  # the reserved slot (col_of_close[x]) IS the spacer
        else:
            col_of[x] = slot
            slot += 1
    return col_of, col_of_close


def sheet_rows(rows: dict[float, int], y0: float, y1: float) -> tuple[int, int]:
    """Inclusive (top_row, bottom_row) worksheet span for normalized
    [y0, y1]. Rows grow downward; y grows upward."""
    top = max(rows.values())
    return (FIRST_GRID_ROW + top - rows[y1], FIRST_GRID_ROW + top - rows[y0] - 1)


def add_schematic_sheet(
    wb: Workbook,
    program: Program,
    theme: Theme,
    *,
    gamma: float = DEFAULT_GAMMA,
    show_premiums: bool = True,
) -> None:
    """Append the tower as a worksheet of merged ranges. The caller owns the
    workbook lifecycle (finalize_workbook runs ONCE, after every sheet)."""
    layout = build_layout(program, gamma=gamma)
    chrome = theme.chrome
    colours = theme.carrier_colours(program.carriers())
    ws = wb.create_sheet(sanitize_sheet_title(f"{program.program} Schematic"))
    ws.sheet_view.showGridLines = False
    _print_setup(ws)

    xs = x_boundaries(layout)
    label_spans = _label_spans(layout, program, show_premiums=show_premiums)
    rows = quantize_boundaries(y_boundaries(layout), label_spans=label_spans)
    top = max(rows.values())
    if len(xs) < 2 or top == 0:  # draft with no drawable tower: title only
        _title(ws, program, chrome, last_col=AXIS_COL)
        return
    col_of, col_of_close = line_column_slots(layout, xs)
    last_col = col_of_close[xs[-1]] - 1
    ws.freeze_panes = f"{get_column_letter(FIRST_GRID_COL)}{FIRST_GRID_ROW}"

    # Fill CANVAS_WIDTH_UNITS regardless of how wide the tower's own layout
    # units happen to span, and regardless of how many line spacers it
    # needs: AXIS_WIDTH and the spacer budget are both fixed off the top,
    # and the tower's own columns share whatever's left, at whatever
    # per-unit rate makes their sum hit the target.
    joined, gapped_starts = _line_transitions(layout)
    n_transitions = max(len(layout.columns) - 1, 0)
    spacer_budget = n_transitions * SPACER_UNITS
    proportional_span = sum(
        x_hi - x_lo for x_lo, x_hi in zip(xs, xs[1:], strict=False)
        if x_lo not in gapped_starts
    )
    chars_per_unit = (CANVAS_WIDTH_UNITS - AXIS_WIDTH - spacer_budget) / proportional_span

    ws.column_dimensions[get_column_letter(AXIS_COL)].width = AXIS_WIDTH
    for x_lo, x_hi in zip(xs, xs[1:], strict=False):
        letter = get_column_letter(col_of[x_lo])
        width = SPACER_UNITS if x_lo in gapped_starts else (x_hi - x_lo) * chars_per_unit
        ws.column_dimensions[letter].width = width
    for x in joined:
        ws.column_dimensions[get_column_letter(col_of_close[x])].width = SPACER_UNITS
    for r in range(FIRST_GRID_ROW, FIRST_GRID_ROW + top):
        ws.row_dimensions[r].height = GRID_ROW_HEIGHT

    _title(ws, program, chrome, last_col=last_col)
    _headers(ws, layout, chrome, col_of, col_of_close, chars_per_unit)
    _axis(ws, layout, rows, chrome)

    pending = {ly.layer_id for ly in layout.layers if labels.is_pending(ly)}
    follows = {ly.id for ly in program.layers if ly.follows_underlying}
    headings = heading_blocks(layout.participants)
    layer_by_id = {ly.layer_id: ly for ly in layout.layers}

    occupied: dict[tuple[int, int], str] = {}

    for index, block in enumerate(layout.participants):
        owner = layer_by_id[block.layer_id]
        is_pending = block.layer_id in pending
        lines = _participant_lines(
            block, owner,
            is_heading=headings.get(block.layer_id) == index,
            follows=block.layer_id in follows,
            is_pending=is_pending,
            show_premiums=show_premiums,
        )
        if block.carrier is None:
            # graphic: pending = empty dashed outline; open remainder = hatch
            fill = (
                None
                if is_pending
                else PatternFill(
                    "lightUp",
                    fgColor=_argb(chrome.unplaced),
                    bgColor=_argb(chrome.background),
                )
            )
            text_colour = chrome.ink if is_pending else chrome.muted
            edge = Side(
                style="dashed" if is_pending else "thin",
                color=_argb(chrome.ink if is_pending else chrome.unplaced),
            )
        else:
            face = colours[block.carrier]
            fill = PatternFill("solid", fgColor=_argb(face))
            text_colour = contrast_text(face, chrome.background, chrome.ink)
            edge = Side(style="thin", color=_argb(chrome.background))
        anchor_rect = max(block.rects, key=lambda r: r.width, default=None)
        if block.carrier is not None and anchor_rect is not None:
            anchor_text, shrink = _fit_label(
                lines, anchor_rect.width * chars_per_unit, block.carrier, block.share_bps
            )
        else:
            anchor_text, shrink = "\n".join(lines), False
        border = Border(left=edge, right=edge, top=edge, bottom=edge)
        if layer_by_id[block.layer_id].statutory:
            # the chevron band IS this bar's top edge; a line here would close
            # the box and read as bounded cover
            border = Border(left=edge, right=edge, bottom=edge)
        for rect in block.rects:
            _block(
                ws, rect, rows, col_of, col_of_close,
                text=anchor_text if rect is anchor_rect else "",
                fill=fill,
                border=border,
                font=Font(name=chrome.font, size=8, color=_argb(text_colour)),
                label=owner.name,
                shrink=shrink if rect is anchor_rect else False,
                occupied=occupied,
            )

    for retention in layout.retentions:
        fill = PatternFill("solid", fgColor=_argb(theme.retention_fill(retention.type)))
        edge = Side(style="thin", color=_argb(chrome.ink))
        for i, rect in enumerate(retention.rects):
            _block(
                ws, rect, rows, col_of, col_of_close,
                text=(
                    retention_label(retention.type, retention.amount, retention.vehicle)
                    if i == 0
                    else ""
                ),
                fill=fill,
                border=Border(left=edge, right=edge, top=edge, bottom=edge),
                font=Font(name=chrome.font, size=7, color=_argb(chrome.ink)),
                label=f"{retention.type} retention",
                occupied=occupied,
            )

    # statutory chevron band: fills its merged width with carets, no fill and
    # no border — it stands in for the bar's top edge, it is not a block
    for band in layout.chevrons:
        width_units = band.width * chars_per_unit
        _block(
            ws, band, rows, col_of, col_of_close,
            text="^" * max(3, int(width_units)),
            fill=None,
            border=Border(),
            font=Font(name=chrome.font, size=8, color=_argb(chrome.ink)),
            label="statutory chevron band",
            occupied=occupied,
        )

    _axis_lines(ws, rows, chrome, last_col)
    _gridlines(ws, layout, rows, chrome, last_col, occupied)
    _footer(ws, chrome, top, last_col)


def _print_setup(ws: Worksheet) -> None:
    """Landscape, fit-to-one-page-wide, comfortably zoomed — a tower reads
    left to right across many lines of cover, so width (not height) is the
    constraint printing needs to respect (Task 11, adopted improvement)."""
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.sheet_view.zoomScale = 80


def _fit_label(
    lines: list[str], width: float, carrier: str, share_bps: int
) -> tuple[str, bool]:
    """Pick the text and alignment for a participant block's anchor cell.
    Below NARROW_MERGE_UNITS_PER_LINE per line, wrap_text degenerates (see
    the module docstring); switch to shrink_to_fit and climb DOWN the
    labels.py compact ladder — this block's full stacked lines, then just
    'Carrier share%', then the carrier name alone, then blank (the fill
    colour still carries the share) — stopping at the first that fits."""

    def fits(candidate: list[str]) -> bool:
        return width >= NARROW_MERGE_UNITS_PER_LINE * max(len(candidate), 1)

    if fits(lines):
        return "\n".join(lines), False
    for candidate in (
        [participant_label(carrier, share_bps)],
        [carrier_only_label(carrier)],
        [],
    ):
        if not candidate or fits(candidate):
            return "\n".join(candidate), True
    return "", True  # unreachable: [] always fits


def _add_border_edge(ws: Worksheet, row: int, col: int, *, side: str, edge: Side) -> None:
    """Add ONE border edge to a cell without disturbing whatever else is
    already there (fill, other edges). The axis lines, gridlines and
    similar post-hoc styling all run AFTER every block/retention merge is
    styled, so a naive `cell.border = Border(...)` would erase the block's
    own edges — this merges the new edge into the existing Border instead
    (same "style after merge, never blindly overwrite" family of rule as
    the merge-before-style fix elsewhere in this module)."""
    cell = ws.cell(row=row, column=col)
    current = cell.border
    sides = {
        "left": current.left, "right": current.right,
        "top": current.top, "bottom": current.bottom,
    }
    sides[side] = edge
    cell.border = Border(**sides)


class SchematicOverlapError(RuntimeError):
    """Two blocks quantized onto the same worksheet cell.

    The tower's rects tile by construction (the module docstring's
    per-BOUNDARY quantization), so this is never a rounding artefact: it
    means the LAYOUT itself has two shapes drawn over each other, which
    only invalid program data produces. The known source is a statutory
    layer — which owns its column floor-to-top, y 0..1, off the dollar
    scale — sharing a line of cover with a dollar-limited layer, exactly
    the `statutory-line-shared` error validate.py already names.

    Refusing here is deliberate. openpyxl's own failure for this is
    `AttributeError: 'MergedCell' object attribute 'value' is read-only`
    from deep inside `_block`, which names neither block nor cell; and
    when the two rects happen to be IDENTICAL (a statutory bar and a
    same-line primary on a single-line program quantize to the same
    range) openpyxl does not fail at all — it silently paints one block
    over the other and ships a wrong client deliverable. Both outcomes
    are worse than a refusal that says which two blocks collided, where."""


def _overlap_guard(
    occupied: dict[tuple[int, int], str],
    label: str,
    r0: int,
    r1: int,
    c0: int,
    c1: int,
) -> None:
    """Refuse BEFORE `ws.merge_cells` mutates anything: a raise from
    half-merged state would leave the caller's workbook inconsistent."""
    for r in range(r0, r1 + 1):
        for c in range(c0, c1 + 1):
            previous = occupied.get((r, c))
            if previous is None:
                continue
            coord = f"{get_column_letter(c)}{r}"
            raise SchematicOverlapError(
                f"schematic blocks overlap at {coord}: {label!r} would draw "
                f"over {previous!r}, which already occupies that cell. Two "
                f"blocks cannot share a cell, so this program's geometry is "
                f"invalid — run `towerctl validate` for the diagnosis (a "
                f"statutory layer owns its whole column, so no other layer "
                f"may cover the same line of cover)."
            )


def _block(
    ws: Worksheet,
    rect: Rect,
    rows: dict[float, int],
    col_of: dict[float, int],
    col_of_close: dict[float, int],
    *,
    text: str,
    fill: PatternFill | None,
    border: Border,
    font: Font,
    label: str,
    shrink: bool = False,
    occupied: dict[tuple[int, int], str] | None = None,
) -> None:
    """One rect → one merged range. MUST merge BEFORE styling non-anchor
    cells: openpyxl's merge_cells() replaces every non-anchor cell with a
    fresh MergedCell (worksheet.py's _clean_merge_range), so any fill/value
    set on those coordinates first is discarded — merge first, style after,
    the same order render_table_sheet already uses in table_xlsx.py.

    `label` names this block for `SchematicOverlapError`; `occupied` both
    feeds `_gridlines` and carries the previous occupant's label, so a
    collision can name BOTH sides."""
    r0, r1 = sheet_rows(rows, rect.y0, rect.y1)
    c0, c1 = col_of[rect.x0], col_of_close[rect.x1] - 1
    if occupied is not None:
        _overlap_guard(occupied, label, r0, r1, c0, c1)
    if (r0, c0) != (r1, c1):  # a 1×1 block is already one cell; no merge
        ws.merge_cells(start_row=r0, start_column=c0, end_row=r1, end_column=c1)
    anchor = ws.cell(row=r0, column=c0, value=text or None)
    anchor.font = font
    anchor.alignment = (
        Alignment(horizontal="center", vertical="center", shrink_to_fit=True)
        if shrink
        else Alignment(horizontal="center", vertical="center", wrap_text=True)
    )
    for r in range(r0, r1 + 1):
        for c in range(c0, c1 + 1):
            cell = ws.cell(row=r, column=c)
            if fill is not None:
                cell.fill = fill
            cell.border = border
            if occupied is not None:
                occupied[(r, c)] = label


def _title(ws: Worksheet, program: Program, chrome: Chrome, *, last_col: int) -> None:
    period = f"{program.period.start.isoformat()} – {program.period.end.isoformat()}"
    cell = ws.cell(
        row=TITLE_ROW, column=AXIS_COL,
        value=f"{program.insured} — {program.program} · {period}",
    )
    cell.font = Font(
        name=chrome.title_font or chrome.font, size=12, color=_argb(chrome.ink)
    )
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if last_col > AXIS_COL:
        ws.merge_cells(
            start_row=TITLE_ROW, start_column=AXIS_COL,
            end_row=TITLE_ROW, end_column=last_col,
        )
    ws.row_dimensions[TITLE_ROW].height = 22.0
    ws.row_dimensions[GROUP_ROW].height = 14.0
    ws.row_dimensions[LINE_ROW].height = 16.0


def _headers(
    ws: Worksheet,
    layout: TowerLayout,
    chrome: Chrome,
    col_of: dict[float, int],
    col_of_close: dict[float, int],
    chars_per_unit: float,
) -> None:
    """Group bands above, line names under them (the spec's header order —
    the graphic puts both below the tower, a chart-only convention).

    Both loops MUST merge before per-cell styling — same rule and same fix
    as `_block` (Task 4's fill loss, Task 9's follow-up: a band spanning 3+
    columns loses its bottom accent border the same way if styled first)."""
    accent = Side(style="medium", color=_argb(chrome.accent))
    for band in layout.groups:
        c0, c1 = col_of[band.x0], col_of_close[band.x1] - 1
        if c1 > c0:
            ws.merge_cells(
                start_row=GROUP_ROW, start_column=c0, end_row=GROUP_ROW, end_column=c1
            )
        cell = ws.cell(row=GROUP_ROW, column=c0, value=group_label(band))
        cell.font = Font(name=chrome.font, size=8, color=_argb(chrome.accent))
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for c in range(c0, c1 + 1):
            ws.cell(row=GROUP_ROW, column=c).border = Border(bottom=accent)

    wraps = False
    for column in layout.columns:
        c0, c1 = col_of[column.ex0], col_of_close[column.ex1] - 1
        if c1 > c0:
            ws.merge_cells(
                start_row=LINE_ROW, start_column=c0, end_row=LINE_ROW, end_column=c1
            )
        cell = ws.cell(row=LINE_ROW, column=c0, value=column.name)
        cell.font = Font(name=chrome.font, size=9, bold=True, color=_argb(chrome.ink))
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        width = (column.ex1 - column.ex0) * chars_per_unit
        if len(column.name) > width:  # never fits on one wrapped line: grow the row
            wraps = True
    if wraps:
        current = ws.row_dimensions[LINE_ROW].height or 0.0
        ws.row_dimensions[LINE_ROW].height = max(current, HEADER_TWO_LINE_HEIGHT)


def _axis(
    ws: Worksheet, layout: TowerLayout, rows: dict[float, int], chrome: Chrome
) -> None:
    """Column A: one merged cell per y-interval, labeled with its floor
    attachment, bottom-aligned — the label sits just above its boundary,
    like the chart's gutter labels. $0 stays silent."""
    ref = list(layout.ref_lines)  # (dollars, y), ascending by construction
    for (d_lo, y_lo), (_d_hi, y_hi) in zip(ref, ref[1:], strict=False):
        if d_lo <= 0:
            continue
        r0, r1 = sheet_rows(rows, y_lo, y_hi)
        cell = ws.cell(row=r0, column=AXIS_COL, value=format_money_compact(d_lo))
        cell.font = Font(name=chrome.font, size=7, color=_argb(chrome.muted))
        cell.alignment = Alignment(horizontal="right", vertical="bottom")
        if r1 > r0:
            ws.merge_cells(
                start_row=r0, start_column=AXIS_COL, end_row=r1, end_column=AXIS_COL
            )


def _axis_lines(
    ws: Worksheet, rows: dict[float, int], chrome: Chrome, last_col: int
) -> None:
    """Real axis lines (Task 11, Grant's Excel review): a solid left border
    down column A's tower extent (Y axis, top boundary to baseline) and a
    solid bottom border along the baseline (y=0) row across the full tower
    width (X axis, every line of cover) — both in the theme's ink colour
    at medium weight. Run AFTER every block/retention merge is styled (see
    `_add_border_edge`), otherwise a block sitting exactly on the baseline
    would overwrite this with its own (thinner, differently-coloured)
    edge.

    Degenerate case: a program with retentions but NO drawable layers has
    y=0.0 AS its own maximum (nothing positive to draw), so `baseline_row`
    lands ABOVE FIRST_GRID_ROW — one row short of colliding with LINE_ROW's
    own header border. Guarded: with no positive-y tower there is no axis
    to draw, so both edges are skipped entirely rather than landing on
    whatever row the arithmetic produces."""
    top = max(rows.values())
    top_row = FIRST_GRID_ROW
    baseline_row = FIRST_GRID_ROW + top - rows[0.0] - 1
    if baseline_row < top_row:
        return
    ink = Side(style="medium", color=_argb(chrome.ink))
    for r in range(top_row, baseline_row + 1):
        _add_border_edge(ws, r, AXIS_COL, side="left", edge=ink)
    for c in range(FIRST_GRID_COL, last_col + 1):
        _add_border_edge(ws, baseline_row, c, side="bottom", edge=ink)


def _gridlines(
    ws: Worksheet,
    layout: TowerLayout,
    rows: dict[float, int],
    chrome: Chrome,
    last_col: int,
    occupied: dict[tuple[int, int], str],
) -> None:
    """Faint hairline top border at every quantized attachment boundary —
    the SAME breakpoints `_axis` labels — but only through EMPTY grid
    cells: a filled participant block already implies the boundary where
    it abuts, so a line straight through one would slice it visually
    instead of marking a boundary (Task 11, adopted improvement). Run
    AFTER the block/retention loop so `occupied` is complete."""
    ref = list(layout.ref_lines)
    hair = Side(style="hair", color=_argb(chrome.grid))
    for (d_lo, y_lo), (_d_hi, y_hi) in zip(ref, ref[1:], strict=False):
        if d_lo <= 0:
            continue
        r0, _ = sheet_rows(rows, y_lo, y_hi)
        for c in range(FIRST_GRID_COL, last_col + 1):
            if (r0, c) not in occupied:
                _add_border_edge(ws, r0, c, side="top", edge=hair)


def _footer(ws: Worksheet, chrome: Chrome, top: int, last_col: int) -> None:
    """A dim, version-only provenance line a few grid rows below the
    tower's lowest drawn row (Task 11, adopted improvement) — mirrors the
    graphic's visible footer text, but deliberately NOT the git sha that
    workbook properties' `provenance()` already carries: a visible cell
    with the sha would change bytes on every commit and break
    SCHEMATIC_GOLDEN_SHA. The version string alone is git-state-independent
    by construction, so the golden hash and two-run byte-identity both stay
    stable across commits."""
    bottom_row = FIRST_GRID_ROW + top - 1
    row = bottom_row + FOOTER_GAP_ROWS
    cell = ws.cell(
        row=row, column=AXIS_COL,
        value=f"towerkit {__version__} · Schedule of Insurance schematic",
    )
    cell.font = Font(name=chrome.font, size=7, color=_argb(chrome.muted))
    cell.alignment = Alignment(horizontal="left", vertical="center")
    end_col = min(AXIS_COL + 5, last_col)
    if end_col > AXIS_COL:
        ws.merge_cells(start_row=row, start_column=AXIS_COL, end_row=row, end_column=end_col)
