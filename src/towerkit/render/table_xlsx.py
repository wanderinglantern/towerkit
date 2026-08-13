"""Generic styled-table workbook writer — the SOI look for any sectioned table.

Extracted from soi_xlsx.py so other tools (bookkit's open-items export)
delegate formatting here instead of copying it. Same determinism contract:
pinned workbook properties + epoch-rewritten archive → byte-identical runs.

Multi-sheet composition contract (PUBLIC API — consumed by this module's
own schematic/SOI sheet writers and by bookkit's future multi-sheet
open-items export): build a workbook (`new_workbook()`, or `Workbook()`
directly), name and render each sheet with
`render_table_sheet` (or another sheet-body writer such as
`render_soi_sheet` / `add_schematic_sheet`) — the first sheet reuses
`wb.active`, later ones come from `wb.create_sheet` — using
`sanitize_sheet_title` for every sheet name, then call
`finalize_workbook` exactly ONCE per workbook after all sheets are
rendered. `write_table` is the single-sheet convenience wrapper over this
same path."""

from __future__ import annotations

import re
import zipfile
from collections.abc import Callable, Sequence
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..theme import Theme
from .common import provenance

_PINNED = datetime(1980, 1, 1)


def _argb(hex_colour: str) -> str:
    return "FF" + hex_colour.lstrip("#").upper()


_PINNED_W3CDTF = b"1980-01-01T00:00:00Z"
_MODIFIED_RE = re.compile(rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)")

_ILLEGAL_SHEET_CHARS = re.compile(r"[\\/?*\[\]:]")
_MULTI_SPACE = re.compile(r" {2,}")


def sanitize_sheet_title(title: str) -> str:
    """openpyxl raises ValueError for `/ \\ ? * [ ] :` in sheet titles and
    caps them at 31 chars. This is the authority for every write_table
    caller — sanitize here so no caller can crash on a client/program name
    that happens to contain one of these (e.g. "A/B Partners"). Illegal
    characters become a space (not dropped, so words don't collide),
    doubled spaces collapse, and the result is trimmed before the cap."""
    cleaned = _MULTI_SPACE.sub(" ", _ILLEGAL_SHEET_CHARS.sub(" ", title)).strip()
    return cleaned[:31].rstrip()


def new_workbook() -> Workbook:
    """Start a multi-sheet composition (see the module docstring's contract).

    Exists so consumers that must not name openpyxl anywhere in their source
    (bookkit convention-tests exactly that, and its strict mypy rejects
    re-importing Workbook from this namespace as an implicit re-export) can
    still obtain the Workbook. Everyone else may construct Workbook directly."""
    return Workbook()


def _normalize_zip(data: bytes, out_path: Path) -> None:
    """Rewrite the archive with epoch timestamps and fixed compression so
    identical content is identical bytes (openpyxl stamps wall-clock zip
    entries). openpyxl's save_workbook() also unconditionally clobbers
    workbook.properties.modified with datetime.now() right before writing
    docProps/core.xml, ignoring our pinned assignment above — so that one
    field has to be neutralized here, post-save, by rewriting its text."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(BytesIO(data)) as src, zipfile.ZipFile(
        out_path, "w", zipfile.ZIP_DEFLATED
    ) as dst:
        for name in src.namelist():
            info = zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            content = src.read(name)
            if name == "docProps/core.xml":
                content = _MODIFIED_RE.sub(
                    lambda m: m.group(1) + _PINNED_W3CDTF + m.group(2), content
                )
            dst.writestr(info, content)


@dataclass(frozen=True)
class TableColumn:
    header: str
    width: float
    number_format: str | None = None
    align: str = "left"   # body-cell horizontal alignment
    wrap: bool = True     # False → wrap_text omitted entirely (see note below)


@dataclass(frozen=True)
class TableSection:
    label: str | None
    rows: tuple[tuple[Any, ...], ...]
    total: Any = None  # rendered in the last column of the label row


def render_table_sheet(
    ws: Worksheet,
    columns: Sequence[TableColumn],
    sections: Sequence[TableSection],
    *,
    theme: Theme,
    row_height: Callable[[tuple[Any, ...]], float] | None = None,
) -> None:
    """The styled sheet body — everything write_table does between naming
    the sheet and pinning workbook properties, verbatim. Title is the
    caller's job so a multi-sheet workbook can name each sheet itself."""
    soi = theme.soi
    ncols = len(columns)
    thin = Side(style="thin", color=_argb(soi.border))
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name=soi.font, size=soi.size, bold=True,
                       color=_argb(soi.effective_header_text))
    body_font = Font(name=soi.font, size=soi.size, color=_argb(soi.body_text))
    header_fill = PatternFill("solid", fgColor=_argb(soi.header_fill))
    band_fill = PatternFill("solid", fgColor=_argb(soi.band_fill))

    for i, col in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = col.width
    for i, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=i, value=col.header)
        cell.font, cell.fill, cell.border = header_font, header_fill, border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 36.0
    ws.freeze_panes = "A2"

    row_ix = 2
    for section in sections:
        if section.label is not None:
            merge_end = ncols - 1 if section.total is not None else ncols
            ws.merge_cells(start_row=row_ix, start_column=1,
                           end_row=row_ix, end_column=merge_end)
            label = ws.cell(row=row_ix, column=1, value=section.label)
            label.font, label.fill = header_font, header_fill
            label.alignment = Alignment(vertical="center")
            if section.total is not None:
                total = ws.cell(row=row_ix, column=ncols, value=section.total)
                total.font, total.fill = header_font, header_fill
                if columns[-1].number_format:
                    total.number_format = columns[-1].number_format
                total.alignment = Alignment(horizontal="right", vertical="center")
            for c in range(1, ncols + 1):
                cell = ws.cell(row=row_ix, column=c)
                cell.fill = header_fill
                cell.border = border
            ws.row_dimensions[row_ix].height = 22.0
            row_ix += 1
        for band, values in enumerate(section.rows):
            for c, (col, value) in enumerate(zip(columns, values, strict=True), start=1):
                cell = ws.cell(row=row_ix, column=c, value=value)
                cell.font, cell.border = body_font, border
                if band % 2 == 1:
                    cell.fill = band_fill
                if col.number_format:
                    cell.number_format = col.number_format
                # BYTE-IDENTICAL SUBTLETY: the old SOI date branch built
                # Alignment(horizontal=..., vertical="top") with NO wrap_text
                # argument. wrap_text=False serializes differently from
                # wrap_text omitted — so wrap=False must OMIT the argument,
                # not pass False.
                if col.wrap:
                    cell.alignment = Alignment(
                        horizontal=col.align, vertical="top", wrap_text=True
                    )
                else:
                    cell.alignment = Alignment(horizontal=col.align, vertical="top")
            if row_height is not None:
                ws.row_dimensions[row_ix].height = row_height(values)
            row_ix += 1


def finalize_workbook(wb: Workbook, out_path: Path) -> Path:
    """Pin properties and normalize the archive — call exactly ONCE per
    workbook, after every sheet is rendered."""
    props = wb.properties
    props.creator = provenance()
    props.created = _PINNED
    props.modified = _PINNED
    props.lastModifiedBy = None
    buffer = BytesIO()
    wb.save(buffer)
    _normalize_zip(buffer.getvalue(), out_path)
    return out_path


def write_table(
    columns: Sequence[TableColumn],
    sections: Sequence[TableSection],
    *,
    title: str,
    theme: Theme,
    out_path: Path,
    row_height: Callable[[tuple[Any, ...]], float] | None = None,
) -> Path:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sanitize_sheet_title(title)
    render_table_sheet(ws, columns, sections, theme=theme, row_height=row_height)
    return finalize_workbook(wb, out_path)
